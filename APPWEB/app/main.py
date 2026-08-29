import datetime
import decimal
import io
import json
import math
import re
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROUND_HALF_UP = decimal.ROUND_HALF_UP

from .db import cursor


app = FastAPI(title="RISEK APPWEB Picking", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


class LoginRequest(BaseModel):
    user_id: int
    password: str
    local_codigo: str
    caja_codigo: int


class SaveRequest(BaseModel):
    data: dict
    original: dict | None = None


class PosLineRequest(BaseModel):
    producto_codigo: str
    cantidad: float
    descuento: float = 0


class PosPaymentRequest(BaseModel):
    fpago_codigo: str
    monto: int
    numero_documento: str | None = None


class PosSaleRequest(BaseModel):
    local_codigo: str
    caja_codigo: int
    user_id: int
    vendedor_codigo: str | None = None
    cliente_rut: str | None = None
    lines: list[PosLineRequest]
    payments: list[PosPaymentRequest]
    validate_only: bool = False


class RoleSaveRequest(BaseModel):
    role_name: str
    is_admin: bool = False
    permissions: list[str] = []


class UserRoleRequest(BaseModel):
    user_id: int
    role_id: int


class CashCloseRequest(BaseModel):
    local_codigo: str
    caja_codigo: int
    user_id: int
    efectivo_contado: int = 0
    observacion: str | None = None


class InvoiceBulkRequest(BaseModel):
    documents: list[dict]
    value: str | None = None


class NvLineRequest(BaseModel):
    producto_codigo: str
    cantidad: float
    descuento: float = 0


class NvSaveRequest(BaseModel):
    local_codigo: str
    user_id: int
    cliente_rut: str
    vendedor_codigo: str
    fecha_emision: str
    fecha_entrega: str
    observacion: str | None = None
    lines: list[NvLineRequest]


MAESTROS = {
    "usuarios": {
        "table": "secuser",
        "title": "Usuarios",
        "key": "SecUserId",
        "name": "SecUserName",
        "order": "SecUserName, SecUserId",
        "fields": [
            ("SecUserId", "Codigo"), ("SecUserName", "Usuario"),
            ("vendedor_codigo", "Vendedor"), ("seccambioprecio", "Cambia precio"),
            ("secusercajero", "Cajero"),
        ],
    },
    "locales": {
        "table": "locales", "title": "Locales", "key": "local_codigo",
        "name": "local_descripcion", "order": "local_descripcion, local_codigo",
        "fields": [
            ("local_codigo", "Codigo"), ("local_descripcion", "Local"),
            ("local_rut", "RUT"), ("local_mail", "Correo"),
            ("local_direccion", "Direccion"), ("local_telefono", "Telefono"),
            ("local_giro", "Giro"), ("local_bodega", "Bodega asociada"),
            ("local_manejocaja", "Maneja caja"), ("local_mailaviso", "Avisos por correo"),
            ("local_name", "Impresora"), ("local_logo_GXI", "Referencia logo"),
            ("local_adress", "Geolocalizacion"), ("local_agenda", "Agenda"),
            ("local_domingo", "Atiende domingo"),
            ("local_iva", "IVA"), ("local_propina", "Propina"),
            ("local_boleta", "Boleta actual"), ("local_factura", "Factura actual"),
            ("local_cotizacion", "Cotizacion actual"), ("local_orden", "Orden actual"),
            ("local_notacredito", "Nota credito inicial"),
            ("local_notacreditofin", "Nota credito final"),
            ("local_dte_inicial", "DTE inicial"), ("local_dte_final", "DTE final"),
            ("local_ge_inicial", "Guia inicial"), ("local_ge_final", "Guia final"),
            ("local_ipservidor", "IP servidor"), ("local_fechaenvio", "Ultimo envio"),
        ],
    },
    "bancos": {
        "table": "bancos", "title": "Bancos", "key": "banco_codigo",
        "name": "banco_descripcion", "order": "banco_descripcion, banco_codigo",
        "fields": [("banco_codigo", "Codigo"), ("banco_descripcion", "Banco")],
    },
    "proveedores": {
        "table": "proveedores", "title": "Proveedores", "key": "proveedor_codigo",
        "name": "proveedor_nombre", "order": "proveedor_nombre, proveedor_codigo",
        "fields": [
            ("proveedor_codigo", "Codigo"), ("proveedor_nombre", "Proveedor"),
            ("proveedor_direccion", "Direccion"), ("proveedor_telefono", "Telefono"),
            ("proveedor_contacto", "Contacto"), ("proveedor_mail", "Correo"),
        ],
    },
    "cajas": {
        "table": "caja", "title": "Cajas", "key": "caja_codigo",
        "name": "local_codigo", "order": "caja_fecha DESC, local_codigo, caja_codigo",
        "fields": [
            ("caja_codigo", "Caja"), ("local_codigo", "Local"),
            ("caja_fecha", "Fecha"), ("caja_estado", "Estado"),
        ],
    },
    "rutas": {
        "table": "rutas", "title": "Rutas", "key": "ruta_id",
        "name": "ruta_nombre", "order": "ruta_nombre, ruta_id",
        "fields": [("ruta_id", "Codigo"), ("ruta_nombre", "Ruta")],
    },
    "bodegas": {
        "table": "bodegas", "title": "Bodegas", "key": "bodega_codigo",
        "name": "bodega_descripcion", "order": "bodega_descripcion, bodega_codigo",
        "fields": [("bodega_codigo", "Codigo"), ("bodega_descripcion", "Bodega")],
    },
    "familias": {
        "table": "familias", "title": "Familias", "key": "familia_codigo",
        "name": "familia_descripcion", "order": "familia_descripcion, familia_codigo",
        "fields": [("familia_codigo", "Codigo"), ("familia_descripcion", "Descripcion")],
    },
    "vendedores": {
        "table": "vendedores", "title": "Vendedores", "key": "vendedor_codigo",
        "name": "vendedor_nombre", "order": "vendedor_nombre, vendedor_codigo",
        "fields": [
            ("vendedor_codigo", "Codigo"), ("vendedor_nombre", "Vendedor"),
            ("vendedor_comision", "Comision (%)"), ("vendedor_estado", "Estado"),
        ],
    },
    "ciudades": {
        "table": "ciudades", "title": "Ciudades", "key": "ciudad_codigo",
        "name": "ciudad_codigo", "order": "ciudad_codigo",
        "fields": [("ciudad_codigo", "Ciudad")],
    },
    "unidades": {
        "table": "unidades", "title": "Unidades", "key": "unidad_codigo",
        "name": "unidad_descripcion", "order": "unidad_descripcion, unidad_codigo",
        "fields": [
            ("unidad_codigo", "Codigo"), ("unidad_descripcion", "Unidad"),
            ("unidad_fconversion", "Factor conversion"),
        ],
    },
    "formasdepago": {
        "table": "formasdepago", "title": "Formas de pago", "key": "fpago_codigo",
        "name": "fpago_descripcion", "order": "fpago_descripcion, fpago_codigo",
        "fields": [
            ("fpago_codigo", "Codigo"), ("fpago_descripcion", "Forma de pago"),
            ("fpago_credito", "Credito"), ("fpago_valor", "Valor"),
            ("fpago_chequeres", "Cheques"), ("fpago_activo", "Activo"),
        ],
    },
    "listaprecios": {
        "table": "listaprecios", "title": "Listas de precios", "key": "lista_codigo",
        "name": "lista_descripcion", "order": "lista_codigo",
        "fields": [("lista_codigo", "Codigo"), ("lista_descripcion", "Descripcion")],
    },
}


def clean_text(value, max_len=None):
    if value is None:
        return None
    if isinstance(value, bytes):
        value = value.decode("latin1", errors="replace")
    text = str(value)
    text = CONTROL_CHARS.sub(" ", text)
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = " ".join(text.split())
    if max_len is not None:
        text = text[:max_len]
    return text


def to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return default
        return int(round(float(value)))
    except Exception:
        return default


def to_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return default
        return out
    except Exception:
        return default


def decimal_value(value, default="0"):
    try:
        if value in (None, ""):
            return decimal.Decimal(default)
        return decimal.Decimal(str(value).replace(",", "."))
    except Exception:
        return decimal.Decimal(default)


def round_business(value):
    return int(decimal_value(value).quantize(decimal.Decimal("1"), rounding=ROUND_HALF_UP))


def val_integer(value):
    match = re.match(r"^[\s]*([+-]?\d+(?:\.\d+)?)", str(value or ""))
    return int(float(match.group(1))) if match else 0


def clean_value(value):
    if isinstance(value, decimal.Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, (bytes, str)):
        return clean_text(value)
    return value


def clean_row(row):
    return {str(k): clean_value(v) for k, v in dict(row).items()}


def clean_rows(rows):
    return [clean_row(r) for r in rows]


def producto_business_values(cur, codigo, incoming, current=None, is_insert=False):
    current = current or {}

    def value(key, default=0):
        raw = incoming.get(key, current.get(key, default))
        return default if raw in (None, "") else raw

    costo = decimal_value(value("producto_costo"))
    margen = decimal_value(value("producto_margenvta"))
    maneja_iva = (clean_text(value("producto_manejaiva", "N"), 1) or "N").upper()
    impuesto_codigo = clean_text(value("impuesto_codigo", ""), 20) or ""
    impuesto_valor = decimal.Decimal("0")
    if impuesto_codigo:
        cur.execute("SELECT COALESCE(impuesto_valor, 0) AS impuesto_valor FROM impuestos WHERE impuesto_codigo = %s LIMIT 1", (impuesto_codigo,))
        impuesto_valor = decimal_value((cur.fetchone() or {}).get("impuesto_valor"))

    neto_base = round_business(costo * (decimal.Decimal("1") + margen / decimal.Decimal("100")))
    iva = round_business(decimal.Decimal(neto_base) * decimal.Decimal("0.19")) if maneja_iva == "S" else 0
    ila = round_business(decimal.Decimal(neto_base) * impuesto_valor / decimal.Decimal("100"))
    venta = round_business(decimal.Decimal(neto_base + iva + ila))
    neto = venta if maneja_iva == "N" else neto_base

    oferta = decimal_value(value("producto_oferta"))
    if maneja_iva == "S":
        divisor = decimal.Decimal("1") + (impuesto_valor + decimal.Decimal("19")) / decimal.Decimal("100")
        oferta_neto = round_business(oferta / divisor) if divisor else 0
    else:
        oferta_neto = venta

    serial = to_int(value("producto_serial"), 0)
    if serial == 0:
        serial = val_integer(codigo)

    calculated = {
        "producto_neto": neto,
        "producto_iva": iva,
        "producto_ila": ila,
        "producto_venta": venta,
        "producto_ofertaneto": oferta_neto,
        "producto_pack": clean_text(value("producto_pack", "N"), 1) or "N",
        "producto_serial": serial,
    }
    if is_insert:
        calculated["producto_estado"] = "2"
    else:
        calculated["producto_ventaant"] = to_int(current.get("producto_venta"), 0)
    return calculated, float(impuesto_valor)


def producto_virtual_values(product, stock=0):
    costo = decimal_value(product.get("producto_costo"))
    margen = decimal_value(product.get("producto_margenvta"))
    stock_value = decimal_value(stock)
    costo_sin_flete = decimal_value(product.get("producto_costosinflete"))
    unidad_envase = decimal_value(product.get("producto_unidadenvase"))
    gramaje = decimal_value(product.get("producto_gramaje"))
    return {
        "producto_netoventa": round_business(
            costo * (decimal.Decimal("1") + margen / decimal.Decimal("100"))
        ),
        "producto_stock": clean_value(stock_value),
        "producto_valorizadocostosin": clean_value(stock_value * costo_sin_flete),
        "producto_stockenvase": clean_value(unidad_envase * gramaje) if unidad_envase > 0 else 0,
    }


def json_safe(data) -> Response:
    return Response(
        content=json.dumps(data, ensure_ascii=False, allow_nan=False, separators=(",", ":")),
        media_type="application/json; charset=utf-8",
    )


def table_columns(cur, table_name: str) -> set[str]:
    try:
        cur.execute(f"SHOW COLUMNS FROM {table_name}")
        return {clean_text(r.get("Field"), 80) for r in clean_rows(cur.fetchall()) if clean_text(r.get("Field"), 80)}
    except Exception:
        return set()


APP_MODULES = [
    ("home", "Dashboard", "Home"), ("punto-venta", "Punto de venta", "Ventas"),
    ("facturas", "Facturas", "Ventas"), ("notas-venta", "Notas de venta", "Ventas"),
    ("guias", "Guias", "Ventas"),
    ("notas-credito", "Notas de credito", "Ventas"), ("boletas", "Boletas", "Ventas"),
    ("picking", "Picking Comercial", "Ventas"), ("usuarios", "Usuarios", "Parametros"),
    ("roles", "Roles y permisos", "Parametros"), ("locales", "Locales", "Parametros"),
    ("bancos", "Bancos", "Parametros"), ("productos", "Productos", "Parametros"),
    ("clientes", "Clientes", "Parametros"), ("vendedores", "Vendedores", "Parametros"),
    ("proveedores", "Proveedores", "Parametros"), ("cajas", "Cajas", "Parametros"),
    ("rutas", "Rutas", "Parametros"), ("ciudades", "Ciudades", "Parametros"),
    ("bodegas", "Bodegas", "Parametros"), ("familias", "Familias", "Parametros"),
    ("unidades", "Unidades", "Parametros"), ("formasdepago", "Formas de pago", "Parametros"),
    ("listaprecios", "Listas de precios", "Parametros"), ("compras", "Ordenes y compras", "Compras"),
    ("recepcion", "Recepcion bodega", "Compras"), ("proveedor-cuenta", "Cuenta proveedor", "Compras"),
    ("reporte-ventas", "Ventas por periodo", "Reporte"), ("reporte-picking", "Picking por ruta", "Reporte"),
    ("reporte-stock", "Stock y productos", "Reporte"), ("reporte-clientes", "Clientes y cobranza", "Reporte"),
    ("rep-estadisticas", "Estadisticas de ventas", "Reporte"), ("rep-vendedores", "Ventas por vendedor", "Reporte"),
    ("rep-rutas", "Ventas por rutas", "Reporte"), ("rep-familias", "Ventas por familias", "Reporte"),
    ("rep-formas-pago", "Ventas por formas de pago", "Reporte"), ("rep-pendientes", "Facturas pendientes", "Reporte"),
    ("rep-cta-cte", "Cuentas corrientes", "Reporte"), ("rep-cobros", "Cobros y recaudacion", "Reporte"),
    ("rep-cartola", "Cartola de clientes", "Reporte"), ("rep-compras", "Compras por proveedor", "Reporte"),
    ("rep-inventario", "Inventario valorizado", "Reporte"), ("rep-stock-bajo", "Stock critico", "Reporte"),
    ("rep-productos", "Ranking de productos", "Reporte"),
    ("gerencia-kpi", "Indicadores KPI", "Gerencia"), ("gerencia-margen", "Margenes", "Gerencia"),
    ("gerencia-rutas", "Rendimiento rutas", "Gerencia"), ("gerencia-caja", "Caja y bancos", "Gerencia"),
]


def ensure_app_security(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS app_roles (
            role_id INT NOT NULL AUTO_INCREMENT,
            role_name VARCHAR(60) NOT NULL,
            is_admin CHAR(1) NOT NULL DEFAULT 'N',
            PRIMARY KEY (role_id), UNIQUE KEY uq_app_role_name (role_name)
        ) ENGINE=InnoDB DEFAULT CHARSET=latin1
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS app_role_permissions (
            role_id INT NOT NULL, module_id VARCHAR(50) NOT NULL,
            PRIMARY KEY (role_id,module_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=latin1
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS app_user_roles (
            user_id INT NOT NULL, role_id INT NOT NULL,
            PRIMARY KEY (user_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=latin1
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS app_caja_cierres (
            local_codigo VARCHAR(10) NOT NULL, caja_codigo INT NOT NULL, cierre_fecha DATE NOT NULL,
            user_id INT NOT NULL, total_ventas BIGINT NOT NULL DEFAULT 0,
            efectivo_sistema BIGINT NOT NULL DEFAULT 0, efectivo_contado BIGINT NOT NULL DEFAULT 0,
            diferencia BIGINT NOT NULL DEFAULT 0, documentos INT NOT NULL DEFAULT 0,
            detalle_pagos TEXT, observacion VARCHAR(250), estado CHAR(1) NOT NULL DEFAULT 'C',
            cierre_hora DATETIME NOT NULL, PRIMARY KEY(local_codigo,caja_codigo,cierre_fecha)
        ) ENGINE=InnoDB DEFAULT CHARSET=latin1
    """)
    cur.execute("INSERT IGNORE INTO app_roles (role_id,role_name,is_admin) VALUES (1,'ADMINISTRADOR','S'),(2,'CAJERO','N')")
    cur.execute("SELECT COUNT(*) AS total FROM app_role_permissions WHERE role_id=2")
    if to_int((cur.fetchone() or {}).get("total"), 0) == 0:
        cur.executemany("INSERT IGNORE INTO app_role_permissions(role_id,module_id) VALUES (2,%s)", [(x,) for x in ("home","punto-venta","boletas")])


def user_access(cur, user_id: int, user_name: str = ""):
    ensure_app_security(cur)
    cur.execute("""
        SELECT r.role_id,r.role_name,r.is_admin
        FROM app_user_roles ur INNER JOIN app_roles r ON r.role_id=ur.role_id
        WHERE ur.user_id=%s LIMIT 1
    """, (user_id,))
    role = clean_row(cur.fetchone() or {})
    if not role:
        admin = user_id == 1 or "ADMIN" in (user_name or "").upper()
        role = {"role_id": 1 if admin else 2, "role_name": "ADMINISTRADOR" if admin else "CAJERO", "is_admin": "S" if admin else "N"}
        cur.execute("INSERT IGNORE INTO app_user_roles(user_id,role_id) VALUES(%s,%s)", (user_id, role["role_id"]))
    if role.get("is_admin") == "S":
        permissions = [module_id for module_id, _, _ in APP_MODULES]
    else:
        cur.execute("SELECT module_id FROM app_role_permissions WHERE role_id=%s", (role.get("role_id"),))
        permissions = [clean_text(row.get("module_id"), 50) for row in cur.fetchall()]
    return role, permissions


def chile_cash_round(value: int) -> int:
    amount = to_int(value, 0)
    remainder = amount % 10
    return amount - remainder if remainder <= 5 else amount + (10 - remainder)


def parse_date(value: str) -> datetime.date:
    try:
        return datetime.date.fromisoformat(clean_text(value, 10) or "")
    except Exception:
        raise HTTPException(400, "Fecha invalida. Use formato YYYY-MM-DD")


def cl_number(value, decimals=2):
    try:
        number = float(value or 0)
    except Exception:
        number = 0.0
    text = f"{number:,.{decimals}f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


def pdf_escape(text) -> str:
    out = clean_text(text, None) or ""
    return out.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text(cmds, x, y, text, size=8, bold=False):
    font = "F2" if bold else "F1"
    cmds.append(f"BT /{font} {size} Tf 0 0 0 rg 1 0 0 1 {x} {y} Tm ({pdf_escape(text)}) Tj ET")


def pdf_text_right(cmds, right_x, y, text, size=8, bold=False):
    text = str(text)
    x = right_x - (len(text) * size * 0.52)
    pdf_text(cmds, x, y, text, size, bold)


def pdf_rect(cmds, x, y, w, h, stroke=True):
    op = "s" if stroke else "f"
    cmds.append(f"0.72 w 0 0 0 RG {x} {y} {w} {h} re {op}")


def pdf_rect_gray(cmds, x, y, w, h, gray=0.82):
    cmds.append(f"0.72 w 0 0 0 RG {gray} {gray} {gray} rg {x} {y} {w} {h} re B")


def pdf_line(cmds, x1, y1, x2, y2, width=0.72):
    cmds.append(f"{width} w 0 0 0 RG {x1} {y1} m {x2} {y2} l S")


def build_pdf(title: str, page_cmds: list[list[str]]) -> bytes:
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        None,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    ]
    kids = []
    for cmds in page_cmds:
        stream = "\n".join(cmds).encode("latin1", errors="replace")
        content_obj = b"<< /Length %d >>\nstream\n%s\nendstream" % (len(stream), stream)
        page_num = len(objects) + 1
        content_num = len(objects) + 2
        kids.append(f"{page_num} 0 R")
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {content_num} 0 R >>".encode(
                "latin1"
            )
        )
        objects.append(content_obj)
    objects[1] = f"<< /Type /Pages /Kids [{' '.join(kids)}] /Count {len(page_cmds)} >>".encode("latin1")

    chunks = [b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n"]
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(sum(len(c) for c in chunks))
        chunks.append(f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n")
    xref_offset = sum(len(c) for c in chunks)
    chunks.append(f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n".encode("ascii"))
    for off in offsets[1:]:
        chunks.append(f"{off:010d} 00000 n \n".encode("ascii"))
    chunks.append(
        f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R /Title ({pdf_escape(title)}) >>\nstartxref\n{xref_offset}\n%%EOF".encode(
            "latin1", errors="replace"
        )
    )
    return b"".join(chunks)


def picking_pdf_bytes(fecha: datetime.date, filtro_label: str, rows: list[dict]) -> bytes:
    emitted = datetime.datetime.now()
    pages = []
    cmds = []
    y = 0
    page_no = 0
    total_kilos = sum(to_float(r.get("kilos"), 0.0) for r in rows)
    total_unidades = sum(to_float(r.get("unidades"), 0.0) for r in rows)

    def new_page():
        nonlocal cmds, y, page_no
        if cmds:
            pages.append(cmds)
        page_no += 1
        cmds = []
        pdf_rect(cmds, 26, 770, 584, 61)
        pdf_text(cmds, 34, 815, "RICARDO SEPULVEDA KURTEN DIST.EIRL (BODEGA)", 8)
        pdf_text(cmds, 34, 800, "RUTA 5 SUR KM 1011 -PTO VARAS", 8)
        pdf_text(cmds, 34, 785, "944056246", 8)
        pdf_text(cmds, 284, 801, "PICKING", 18)
        pdf_text(cmds, 365, 802, fecha.strftime("%d/%m/%y"), 15)
        pdf_text(cmds, 500, 815, "Fecha", 8)
        pdf_text(cmds, 532, 815, emitted.strftime("%d/%m/%y"), 8)
        pdf_text(cmds, 500, 797, "Hora", 8)
        pdf_text(cmds, 532, 797, emitted.strftime("%H:%M:%S"), 8)
        if filtro_label and filtro_label != "Todas las rutas":
            pdf_text(cmds, 318, 781, filtro_label.upper(), 7)
        if page_no == 1:
            pdf_rect_gray(cmds, 26, 730, 584, 22)
            pdf_text(cmds, 34, 738, "CODIGO", 8)
            pdf_text(cmds, 115, 738, "DESCRIPCION", 8)
            pdf_text(cmds, 393, 738, "CAJAS", 8)
            pdf_text(cmds, 451, 738, "RESTO", 8)
            pdf_text(cmds, 509, 738, "KILOS", 8)
            pdf_text(cmds, 551, 738, "UN.VENDIDAS", 8)
            y = 714
        else:
            y = 717

    new_page()
    for row in rows:
        if y < 74:
            new_page()
        pdf_text(cmds, 26, y, clean_text(row.get("producto_codigo"), 14) or "-", 8)
        pdf_text(cmds, 101, y, (clean_text(row.get("descripcion"), 56) or "-")[:56], 8)
        cajas = to_float(row.get("cajas"), 0.0)
        if abs(cajas) > 0.0001:
            pdf_text_right(cmds, 432, y, cl_number(cajas, 3), 8)
        resto = to_float(row.get("resto"), 0.0)
        if abs(resto) > 0.0001:
            pdf_text_right(cmds, 486, y, cl_number(resto, 0), 8)
        pdf_text_right(cmds, 542, y, cl_number(row.get("kilos"), 2), 8)
        pdf_text_right(cmds, 606, y, cl_number(row.get("unidades"), 2), 8)
        pdf_line(cmds, 26, y - 5, 607, y - 5)
        for x in (380, 434, 488):
            pdf_rect(cmds, x, y - 4, 1.4, 13.5)
        y -= 16

    if y < 216:
        new_page()
    pdf_line(cmds, 26, y + 3, 607, y + 3, 1.44)
    pdf_line(cmds, 26, y - 1, 607, y - 1, 0.72)
    pdf_text(cmds, 219, y - 15, "TOTAL KILOS", 8)
    pdf_text_right(cmds, 542, y - 15, cl_number(total_kilos, 2), 8)
    pdf_text_right(cmds, 606, y - 15, cl_number(total_unidades, 2), 8)
    money_y = y - 178
    pdf_rect(cmds, 24, money_y, 588, 145)
    pdf_text(cmds, 290, money_y + 118, "DESGLOSE DE DINERO", 12, True)
    for idx, label in enumerate(("EFECTIVO", "CHEQUES:", "CREDITOS", "GASTOS", "TRANSFERENCIA:", "TOTAL :")):
        yy = money_y + 100 - (idx * 18)
        pdf_text_right(cmds, 108, yy, label, 8)
        pdf_line(cmds, 119, yy - 1, 192, yy - 1)
    bills_left = (("20.000", 95), ("10.000", 72), ("5.000", 49), ("2.000", 26))
    bills_right = (("1.000", 95), ("500", 72), ("100", 49), ("50", 26), ("10", 10))
    for label, offset in bills_left:
        yy = money_y + offset
        pdf_text(cmds, 286, yy, label, 8)
        pdf_line(cmds, 316, yy - 1, 389, yy - 1)
    for label, offset in bills_right:
        yy = money_y + offset
        pdf_text(cmds, 403, yy, label, 8)
        pdf_line(cmds, 433, yy - 1, 506, yy - 1)
    pages.append(cmds)
    return build_pdf("Picking", pages)


def picking_data(fecha: datetime.date, ruta_id: str | None, vendedor_codigo: str | None):
    start = fecha
    end = fecha + datetime.timedelta(days=1)
    params = [start, end]
    filters = []
    if ruta_id:
        filters.append("AND CAST(c.ruta_id AS CHAR) = %s")
        params.append(ruta_id)
    if vendedor_codigo:
        filters.append("AND CAST(v.vendedor_codigo AS CHAR) = %s")
        params.append(vendedor_codigo)
    extra_filter = "\n".join(filters)

    with cursor() as (_, cur):
        cur.execute(f"""
            SELECT
                vl.producto_codigo,
                COALESCE(NULLIF(vl.venta_descripcion,''), p.producto_descripcion, vl.producto_codigo) AS descripcion,
                COALESCE(p.producto_unidadenvase, 0) AS unidad_envase,
                COALESCE(SUM(COALESCE(vl.venta_unidadenvase,0)),0) AS unidades,
                COALESCE(SUM(COALESCE(vl.venta_kilos,0)),0) AS kilos,
                COUNT(DISTINCT CONCAT(v.venta_tipo, '-', v.local_codigo, '-', v.venta_numero)) AS documentos
            FROM ventas v
            INNER JOIN ventaslevel2 vl
              ON vl.venta_numero = v.venta_numero
             AND vl.venta_tipo = v.venta_tipo
             AND vl.local_codigo = v.local_codigo
            LEFT JOIN productos p ON p.producto_codigo = vl.producto_codigo
            LEFT JOIN clientes c ON c.cliente_rut = v.cliente_rut
            WHERE v.venta_tipo IN ('FE','FA','BO')
              AND COALESCE(v.venta_numero,0) > 0
              AND v.venta_fecha >= %s
              AND v.venta_fecha < %s
              {extra_filter}
            GROUP BY vl.producto_codigo, COALESCE(NULLIF(vl.venta_descripcion,''), p.producto_descripcion, vl.producto_codigo), COALESCE(p.producto_unidadenvase, 0)
            ORDER BY descripcion, vl.producto_codigo
        """, tuple(params))
        rows = []
        for r in clean_rows(cur.fetchall()):
            unidades = to_float(r.get("unidades"), 0.0)
            unidad_envase = to_float(r.get("unidad_envase"), 0.0)
            if unidad_envase > 0:
                cajas = math.floor(unidades / unidad_envase)
                resto = unidades - (cajas * unidad_envase)
            else:
                cajas = 0
                resto = unidades
            rows.append({
                "producto_codigo": clean_text(r.get("producto_codigo"), 20) or "-",
                "descripcion": clean_text(r.get("descripcion"), 80) or "-",
                "unidad_envase": unidad_envase,
                "cajas": cajas,
                "resto": resto,
                "kilos": to_float(r.get("kilos"), 0.0),
                "unidades": unidades,
                "documentos": to_int(r.get("documentos"), 0),
            })
        return rows


@app.get("/")
def index():
    return FileResponse(str(STATIC_DIR / "index.html"))


@app.get("/api/health")
def health():
    return {"ok": True, "service": "risek-appweb-picking"}


@app.get("/api/dashboard")
def dashboard(local_codigo: str | None = Query(None)):
    today = datetime.date.today()
    since = today - datetime.timedelta(days=6)
    selected_local = clean_text(local_codigo, 10) or ""
    local_clause = " AND local_codigo = %s" if selected_local else ""
    venta_params = lambda *values: tuple(values) + ((selected_local,) if selected_local else ())
    metrics = {
        "ventas_dia": 0,
        "nv_dia": 0,
        "facturas_entregar": 0,
        "rutas_entrega": 0,
        "compras_7_dias": 0,
        "clientes_activos": 0,
    }
    queries = {
        "ventas_dia": (f"SELECT COUNT(*) AS total FROM ventas WHERE venta_fecha = %s{local_clause}", venta_params(today)),
        "nv_dia": (f"SELECT COUNT(*) AS total FROM ventas WHERE venta_fecha = %s AND COALESCE(venta_nv, 0) > 0{local_clause}", venta_params(today)),
        "facturas_entregar": ("""
            SELECT COUNT(*) AS total
            FROM ventas
            WHERE venta_fecha = %s
              AND venta_tipo IN ('FE', 'FA', 'GD')
              AND COALESCE(venta_entregado, 'N') NOT IN ('S', '1')
        """ + local_clause, venta_params(today)),
        "rutas_entrega": ("""
            SELECT COUNT(DISTINCT c.ruta_id) AS total
            FROM ventas v
            INNER JOIN clientes c ON c.cliente_rut = v.cliente_rut
            WHERE v.venta_fecha = %s
              AND v.venta_tipo IN ('FE', 'FA', 'GD')
              AND COALESCE(v.venta_entregado, 'N') NOT IN ('S', '1')
              AND c.ruta_id IS NOT NULL
        """ + (" AND v.local_codigo = %s" if selected_local else ""), venta_params(today)),
        "compras_7_dias": (f"SELECT COUNT(*) AS total FROM compras WHERE compra_fecha BETWEEN %s AND %s{local_clause}", venta_params(since, today)),
        "clientes_activos": ("SELECT COUNT(*) AS total FROM clientes WHERE COALESCE(cliente_estado, 'A') NOT IN ('I', 'N')", ()),
    }
    with cursor() as (_, cur):
        for key, (sql, params) in queries.items():
            try:
                cur.execute(sql, params)
                metrics[key] = to_int((cur.fetchone() or {}).get("total"), 0)
            except Exception:
                metrics[key] = 0
        metrics["local"] = None
        metrics["folios"] = {"facturas": 0, "boletas": 0, "guias": 0, "notas_credito": 0}
        if selected_local:
            try:
                cur.execute("SELECT * FROM locales WHERE local_codigo = %s LIMIT 1", (selected_local,))
                local = clean_row(cur.fetchone() or {})
                if local:
                    metrics["local"] = {
                        "codigo": clean_text(local.get("local_codigo"), 10) or selected_local,
                        "descripcion": clean_text(local.get("local_descripcion"), 100) or selected_local,
                        "bodega": clean_text(local.get("local_bodega"), 30) or "",
                    }
            except Exception:
                pass
        try:
            cur.execute("SELECT * FROM locales WHERE local_codigo IN ('01', '02')")
            folio_locales = {
                clean_text(row.get("local_codigo"), 10): clean_row(row)
                for row in cur.fetchall()
            }
            local_01 = folio_locales.get("01", {})
            local_02 = folio_locales.get("02", {})
            metrics["folios"] = {
                "facturas": to_int(local_01.get("local_dte_final"), 0) - to_int(local_01.get("local_dte_inicial"), 0),
                "boletas": to_int(local_02.get("local_factura"), 0) - to_int(local_02.get("local_boleta"), 0),
                "guias": to_int(local_01.get("local_ge_final"), 0) - to_int(local_01.get("local_ge_inicial"), 0),
                "notas_credito": to_int(local_01.get("local_notacreditofin"), 0) - to_int(local_01.get("local_notacredito"), 0),
            }
        except Exception:
            pass
    metrics["fecha"] = today.isoformat()
    return json_safe(metrics)


@app.get("/api/dashboard/charts")
def dashboard_charts(local_codigo: str | None = Query(None)):
    today = datetime.date.today()
    current_year = today.year
    previous_year = current_year - 1
    annual_start = datetime.date(previous_year, 1, 1)
    annual_end = datetime.date(current_year + 1, 1, 1)
    month_start = today.replace(day=1)
    last_month_start = (month_start - datetime.timedelta(days=1)).replace(day=1)
    selected_local = clean_text(local_codigo, 10) or ""
    local_clause = " AND local_codigo = %s" if selected_local else ""
    with_local = lambda *values: tuple(values) + ((selected_local,) if selected_local else ())

    with cursor() as (_, cur):
        cur.execute("""
            SELECT venta_fecha AS fecha, COUNT(*) AS total
            FROM ventas
            WHERE venta_tipo = 'NV' AND venta_fecha >= %s AND venta_fecha < %s
            """ + local_clause + """
            GROUP BY venta_fecha ORDER BY venta_fecha
        """, with_local(last_month_start, month_start))
        pedidos = clean_rows(cur.fetchall())
        cur.execute("""
            SELECT venta_fecha AS fecha, COUNT(*) AS total
            FROM ventas
            WHERE venta_tipo IN ('FE','FA') AND venta_fecha >= %s AND venta_fecha <= %s
            """ + local_clause + """
            GROUP BY venta_fecha ORDER BY venta_fecha
        """, with_local(month_start, today))
        facturas_dia = clean_rows(cur.fetchall())
        cur.execute("""
            SELECT YEAR(venta_fecha) AS ano, MONTH(venta_fecha) AS mes, COUNT(*) AS total
            FROM ventas
            WHERE venta_tipo IN ('FE','FA') AND venta_fecha >= %s AND venta_fecha < %s
            """ + local_clause + """
            GROUP BY YEAR(venta_fecha), MONTH(venta_fecha)
            ORDER BY ano, mes
        """, with_local(annual_start, annual_end))
        facturas_ano = clean_rows(cur.fetchall())
    return json_safe({
        "pedidos_mes_anterior": pedidos,
        "facturas_mes_actual": facturas_dia,
        "facturas_por_mes": facturas_ano,
        "current_year": current_year,
        "previous_year": previous_year,
    })


@app.get("/api/secusers")
def secusers():
    with cursor() as (_, cur):
        try:
            cur.execute("""
                SELECT
                    SecUserId AS user_id,
                    COALESCE(NULLIF(SecUserName,''), CONCAT('Usuario ', SecUserId)) AS user_name,
                    vendedor_codigo,
                    COALESCE(seccambioprecio, 'N') AS cambia_precio,
                    COALESCE(secusercajero, 'N') AS cajero
                FROM secuser
                ORDER BY SecUserName, SecUserId
            """)
            return json_safe(clean_rows(cur.fetchall()))
        except Exception:
            return json_safe([])


@app.get("/api/locales/login")
def login_locales():
    with cursor() as (_, cur):
        try:
            cur.execute("""
                SELECT local_codigo, local_descripcion, local_bodega
                FROM locales
                ORDER BY local_descripcion, local_codigo
            """)
            return json_safe(clean_rows(cur.fetchall()))
        except Exception:
            return json_safe([])


@app.get("/api/cajas/login")
def login_cajas():
    with cursor() as (_, cur):
        cur.execute("""
            SELECT c.local_codigo, c.caja_codigo, c.caja_fecha, COALESCE(c.caja_estado, '') AS caja_estado
            FROM caja c
            INNER JOIN (
                SELECT local_codigo, caja_codigo, MAX(caja_fecha) AS caja_fecha
                FROM caja
                WHERE COALESCE(local_codigo, '') <> '' AND caja_codigo > 0
                GROUP BY local_codigo, caja_codigo
            ) last_caja
              ON last_caja.local_codigo = c.local_codigo
             AND last_caja.caja_codigo = c.caja_codigo
             AND last_caja.caja_fecha = c.caja_fecha
            ORDER BY c.local_codigo, c.caja_codigo
        """)
        return json_safe(clean_rows(cur.fetchall()))


@app.post("/api/login")
def login(payload: LoginRequest):
    user_id = to_int(payload.user_id, 0)
    password = clean_text(payload.password, 100) or ""
    local_codigo = clean_text(payload.local_codigo, 10) or ""
    caja_codigo = to_int(payload.caja_codigo, 0)
    if user_id <= 0:
        raise HTTPException(400, "Seleccione un usuario")
    if not local_codigo:
        raise HTTPException(400, "Seleccione un local")
    if caja_codigo <= 0:
        raise HTTPException(400, "Seleccione una caja")

    with cursor() as (_, cur):
        cur.execute("""
            SELECT
                SecUserId AS user_id,
                COALESCE(NULLIF(SecUserName,''), CONCAT('Usuario ', SecUserId)) AS user_name,
                COALESCE(SecUserPassword, '') AS user_password,
                vendedor_codigo,
                COALESCE(seccambioprecio, 'N') AS cambia_precio,
                COALESCE(secusercajero, 'N') AS cajero
            FROM secuser
            WHERE SecUserId = %s
            LIMIT 1
        """, (user_id,))
        row = cur.fetchone()
        cur.execute("""
            SELECT local_codigo, local_descripcion, local_bodega
            FROM locales WHERE local_codigo = %s LIMIT 1
        """, (local_codigo,))
        local = cur.fetchone()
        cur.execute("""
            SELECT local_codigo, caja_codigo, MAX(caja_fecha) AS caja_fecha
            FROM caja
            WHERE local_codigo = %s AND caja_codigo = %s
            GROUP BY local_codigo, caja_codigo
        """, (local_codigo, caja_codigo))
        caja = cur.fetchone()

    if not row:
        raise HTTPException(401, "Usuario no encontrado")
    if not local:
        raise HTTPException(400, "Local no encontrado")
    if not caja:
        raise HTTPException(400, "La caja no pertenece al local seleccionado")

    user = clean_row(row)
    stored_password = clean_text(user.pop("user_password", ""), 100) or ""
    if password != stored_password:
        raise HTTPException(401, "Contrasena incorrecta")
    with cursor() as (_, cur):
        role, permissions = user_access(cur, user_id, clean_text(user.get("user_name"), 100) or "")
    user["role"] = role
    user["permissions"] = permissions

    return json_safe({
        "ok": True,
        "user": user,
        "local": clean_row(local),
        "caja": clean_row(caja),
    })


def pos_price_list(local_codigo: str) -> str:
    return "30" if local_codigo == "02" else "01"


@app.get("/api/roles")
def roles_listado():
    with cursor() as (_, cur):
        ensure_app_security(cur)
        cur.execute("SELECT role_id,role_name,is_admin FROM app_roles ORDER BY role_name")
        roles = clean_rows(cur.fetchall())
        for role in roles:
            cur.execute("SELECT module_id FROM app_role_permissions WHERE role_id=%s", (role["role_id"],))
            role["permissions"] = [row["module_id"] for row in cur.fetchall()]
        users = []
        if "SecUserId" in table_columns(cur, "secuser"):
            cur.execute("""
                SELECT u.SecUserId AS user_id,u.SecUserName AS user_name,COALESCE(ur.role_id,0) AS role_id
                FROM secuser u LEFT JOIN app_user_roles ur ON ur.user_id=u.SecUserId
                ORDER BY u.SecUserName
            """)
            users = clean_rows(cur.fetchall())
    return json_safe({"roles": roles, "users": users, "modules": [{"id":x,"label":y,"group":z} for x,y,z in APP_MODULES]})


@app.post("/api/roles")
def role_guardar(payload: RoleSaveRequest):
    name = clean_text(payload.role_name, 60) or ""
    if not name:
        raise HTTPException(400, "Ingrese nombre del rol")
    with cursor() as (_, cur):
        ensure_app_security(cur)
        cur.execute("INSERT INTO app_roles(role_name,is_admin) VALUES(%s,%s)", (name, "S" if payload.is_admin else "N"))
        role_id = cur.lastrowid
        if not payload.is_admin:
            cur.executemany("INSERT IGNORE INTO app_role_permissions(role_id,module_id) VALUES(%s,%s)", [(role_id, clean_text(x,50)) for x in payload.permissions])
    return json_safe({"ok": True, "role_id": role_id})


@app.put("/api/roles/{role_id}")
def role_actualizar(role_id: int, payload: RoleSaveRequest):
    if role_id == 1 and not payload.is_admin:
        raise HTTPException(400, "El rol administrador debe conservar acceso total")
    with cursor() as (_, cur):
        ensure_app_security(cur)
        cur.execute("UPDATE app_roles SET role_name=%s,is_admin=%s WHERE role_id=%s", (clean_text(payload.role_name,60), "S" if payload.is_admin else "N", role_id))
        cur.execute("DELETE FROM app_role_permissions WHERE role_id=%s", (role_id,))
        if not payload.is_admin:
            cur.executemany("INSERT IGNORE INTO app_role_permissions(role_id,module_id) VALUES(%s,%s)", [(role_id, clean_text(x,50)) for x in payload.permissions])
    return json_safe({"ok": True})


@app.put("/api/roles/usuario/asignar")
def usuario_asignar_rol(payload: UserRoleRequest):
    with cursor() as (_, cur):
        ensure_app_security(cur)
        cur.execute("SELECT 1 FROM app_roles WHERE role_id=%s", (payload.role_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Rol no encontrado")
        cur.execute("INSERT INTO app_user_roles(user_id,role_id) VALUES(%s,%s) ON DUPLICATE KEY UPDATE role_id=VALUES(role_id)", (payload.user_id,payload.role_id))
    return json_safe({"ok": True})


@app.get("/api/pos/caja/status")
def pos_caja_status(local_codigo: str = Query(...), caja_codigo: int = Query(...)):
    local = clean_text(local_codigo,10) or ""
    today = datetime.date.today()
    with cursor() as (_,cur):
        ensure_app_security(cur)
        if not {"ventas", "ventaslevel1"}.issubset({"ventas" if table_columns(cur,"ventas") else "", "ventaslevel1" if table_columns(cur,"ventaslevel1") else ""}):
            return json_safe({"cerrada": False, "disponible": False, "mensaje": "Tablas de ventas no disponibles", "documentos":0, "total":0, "pagos":[]})
        cur.execute("SELECT * FROM app_caja_cierres WHERE local_codigo=%s AND caja_codigo=%s AND cierre_fecha=%s", (local,caja_codigo,today))
        close = clean_row(cur.fetchone() or {})
        cur.execute("SELECT COUNT(*) documentos,COALESCE(SUM(venta_totalventa),0) total FROM ventas WHERE venta_tipo='BO' AND local_codigo=%s AND caja_codigo=%s AND venta_fecha=%s", (local,caja_codigo,today))
        totals = clean_row(cur.fetchone() or {})
        cur.execute("""
            SELECT p.fpago_codigo,COALESCE(fp.fpago_descripcion,p.fpago_codigo) descripcion,COALESCE(SUM(p.venta_pagomonto),0) monto
            FROM ventaslevel1 p INNER JOIN ventas v ON v.venta_numero=p.venta_numero AND v.venta_tipo=p.venta_tipo AND v.local_codigo=p.local_codigo
            LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo
            WHERE v.venta_tipo='BO' AND v.local_codigo=%s AND v.caja_codigo=%s AND v.venta_fecha=%s
            GROUP BY p.fpago_codigo,fp.fpago_descripcion ORDER BY descripcion
        """, (local,caja_codigo,today))
        payments = clean_rows(cur.fetchall())
    return json_safe({"cerrada": bool(close), "cierre": close, "documentos":to_int(totals.get("documentos"),0), "total":to_int(totals.get("total"),0), "pagos":payments})


@app.post("/api/pos/caja/cerrar")
def pos_cerrar_caja(payload: CashCloseRequest):
    local=clean_text(payload.local_codigo,10) or ""; today=datetime.date.today()
    with cursor() as (_,cur):
        ensure_app_security(cur)
        if not table_columns(cur,"ventas") or not table_columns(cur,"ventaslevel1"):
            raise HTTPException(503, "Las tablas de ventas no estan disponibles")
        cur.execute("SELECT COUNT(*) documentos,COALESCE(SUM(venta_totalventa),0) total FROM ventas WHERE venta_tipo='BO' AND local_codigo=%s AND caja_codigo=%s AND venta_fecha=%s",(local,payload.caja_codigo,today))
        totals=clean_row(cur.fetchone() or {})
        cur.execute("""SELECT p.fpago_codigo,COALESCE(fp.fpago_descripcion,p.fpago_codigo) descripcion,COALESCE(SUM(p.venta_pagomonto),0) monto FROM ventaslevel1 p INNER JOIN ventas v ON v.venta_numero=p.venta_numero AND v.venta_tipo=p.venta_tipo AND v.local_codigo=p.local_codigo LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo WHERE v.venta_tipo='BO' AND v.local_codigo=%s AND v.caja_codigo=%s AND v.venta_fecha=%s GROUP BY p.fpago_codigo,fp.fpago_descripcion""",(local,payload.caja_codigo,today))
        payments=clean_rows(cur.fetchall()); cash=sum(to_int(x.get("monto"),0) for x in payments if clean_text(x.get("fpago_codigo"),10)=="01")
        counted=to_int(payload.efectivo_contado,0)
        cur.execute("""INSERT INTO app_caja_cierres(local_codigo,caja_codigo,cierre_fecha,user_id,total_ventas,efectivo_sistema,efectivo_contado,diferencia,documentos,detalle_pagos,observacion,estado,cierre_hora) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'C',%s) ON DUPLICATE KEY UPDATE user_id=VALUES(user_id),total_ventas=VALUES(total_ventas),efectivo_sistema=VALUES(efectivo_sistema),efectivo_contado=VALUES(efectivo_contado),diferencia=VALUES(diferencia),documentos=VALUES(documentos),detalle_pagos=VALUES(detalle_pagos),observacion=VALUES(observacion),estado='C',cierre_hora=VALUES(cierre_hora)""",(local,payload.caja_codigo,today,payload.user_id,to_int(totals.get("total"),0),cash,counted,counted-cash,to_int(totals.get("documentos"),0),json.dumps(payments,ensure_ascii=False),clean_text(payload.observacion,250),datetime.datetime.now()))
        cur.execute("UPDATE caja SET caja_estado='C',caja_ventas=%s WHERE local_codigo=%s AND caja_codigo=%s AND caja_fecha=(SELECT max_fecha FROM (SELECT MAX(caja_fecha) max_fecha FROM caja WHERE local_codigo=%s AND caja_codigo=%s) x)",(to_int(totals.get("total"),0),local,payload.caja_codigo,local,payload.caja_codigo))
    return json_safe({"ok":True,"total":to_int(totals.get("total"),0),"efectivo_sistema":cash,"diferencia":counted-cash})


def pos_product(cur, scanned_code: str, local_codigo: str):
    scanned = clean_text(scanned_code, 30) or ""
    if not scanned:
        return None
    quantity = decimal.Decimal("1")
    source = "codigo"
    cur.execute("SELECT producto_codigo FROM productos WHERE producto_codigo = %s LIMIT 1", (scanned,))
    found = cur.fetchone()
    product_code = clean_text((found or {}).get("producto_codigo"), 20)
    if not product_code:
        cur.execute("SELECT producto_codigo FROM productos WHERE producto_barra = %s LIMIT 1", (scanned,))
        found = cur.fetchone()
        product_code = clean_text((found or {}).get("producto_codigo"), 20)
        source = "barra"
    if not product_code and len(scanned) >= 12 and scanned.isdigit():
        product_code = str(to_int(scanned[2:7], 0))
        quantity = decimal_value(scanned[7:12])
        source = "barra_peso"
    if not product_code:
        return None

    price_list = pos_price_list(local_codigo)
    cur.execute("""
        SELECT p.producto_codigo, p.producto_descripcion, p.producto_barra,
               p.unidad_codigo, p.producto_unidadenvase, p.producto_descuentastock,
               COALESCE(pl.lista_neto, p.producto_neto, 0) AS precio_neto,
               COALESCE(pl.lista_iva, p.producto_iva, 0) AS precio_iva,
               COALESCE(pl.lista_ila, p.producto_ila, 0) AS precio_ila,
               COALESCE(pl.lista_venta, p.producto_venta, 0) AS precio_venta
        FROM productos p
        LEFT JOIN precioslevel1 pl
          ON pl.producto_codigo = p.producto_codigo AND pl.lista_codigo = %s
        WHERE p.producto_codigo = %s
          AND COALESCE(p.producto_estado, 'A') NOT IN ('I', 'N')
        LIMIT 1
    """, (price_list, product_code))
    product = clean_row(cur.fetchone() or {})
    if not product:
        return None
    if source == "barra_peso" and (clean_text(product.get("unidad_codigo"), 3) or "").upper() == "KG":
        quantity = (quantity / decimal.Decimal("1000")).quantize(decimal.Decimal("0.001"), rounding=ROUND_HALF_UP)
    product.update({
        "cantidad": clean_value(quantity),
        "origen": source,
        "lista_codigo": price_list,
        "codigo_escaneado": scanned,
    })
    return product


@app.get("/api/pos/catalogos")
def pos_catalogos(local_codigo: str = Query(...)):
    local = clean_text(local_codigo, 10) or ""
    with cursor() as (_, cur):
        cur.execute("""
            SELECT fpago_codigo AS value, COALESCE(fpago_descripcion, fpago_codigo) AS label
            FROM formasdepago
            WHERE COALESCE(fpago_activo, 'S') <> 'N'
            ORDER BY fpago_descripcion
        """)
        payments = clean_rows(cur.fetchall())
        cur.execute("SELECT vendedor_codigo AS value, vendedor_nombre AS label FROM vendedores WHERE COALESCE(vendedor_estado, 'A') <> 'B' ORDER BY vendedor_nombre")
        sellers = clean_rows(cur.fetchall())
        cur.execute("SELECT local_boleta, local_bodega FROM locales WHERE local_codigo = %s LIMIT 1", (local,))
        local_row = clean_row(cur.fetchone() or {})
    return json_safe({
        "formas_pago": payments,
        "vendedores": sellers,
        "lista_codigo": pos_price_list(local),
        "folio_siguiente": to_int(local_row.get("local_boleta"), 0),
        "bodega_codigo": clean_text(local_row.get("local_bodega"), 10) or local,
    })


@app.get("/api/boletas")
def boletas_listado(
    local_codigo: str | None = Query(None),
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    numero: str | None = Query(None),
    vendedor_codigo: str | None = Query(None),
    fpago_codigo: str | None = Query(None),
    limit: int = Query(200, ge=1, le=500),
):
    filters = ["v.venta_tipo = 'BO'"]
    params = []
    local = clean_text(local_codigo, 10)
    if local:
        filters.append("v.local_codigo = %s")
        params.append(local)
    if fecha_desde:
        filters.append("v.venta_fecha >= %s")
        params.append(parse_date(fecha_desde))
    if fecha_hasta:
        filters.append("v.venta_fecha <= %s")
        params.append(parse_date(fecha_hasta))
    number = clean_text(numero, 20)
    if number:
        filters.append("CAST(v.venta_numero AS CHAR) LIKE %s")
        params.append(f"{number}%")
    seller = clean_text(vendedor_codigo, 10)
    if seller:
        filters.append("v.vendedor_codigo = %s")
        params.append(seller)
    payment = clean_text(fpago_codigo, 10)
    if payment:
        filters.append("EXISTS (SELECT 1 FROM ventaslevel1 vp WHERE vp.venta_numero=v.venta_numero AND vp.venta_tipo=v.venta_tipo AND vp.local_codigo=v.local_codigo AND vp.fpago_codigo=%s)")
        params.append(payment)
    where_sql = " AND ".join(filters)
    with cursor() as (_, cur):
        cur.execute(f"""
            SELECT v.venta_numero, v.local_codigo, v.venta_fecha, v.venta_hora,
                   v.cliente_rut, COALESCE(c.cliente_nombre, 'CONSUMIDOR FINAL') AS cliente_nombre,
                   v.vendedor_codigo, COALESCE(ve.vendedor_nombre, v.vendedor_codigo) AS vendedor_nombre,
                   v.caja_codigo, v.venta_estado, v.venta_totalventa, v.venta_neto1,
                   v.venta_iva1, v.venta_ila1, v.venta_pagototal, v.venta_pago,
                   (SELECT COUNT(*) FROM ventaslevel2 d WHERE d.venta_numero=v.venta_numero AND d.venta_tipo=v.venta_tipo AND d.local_codigo=v.local_codigo) AS lineas,
                   (SELECT GROUP_CONCAT(DISTINCT COALESCE(fp.fpago_descripcion,p.fpago_codigo) ORDER BY p.venta_pagoitem SEPARATOR ' + ')
                    FROM ventaslevel1 p LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo
                    WHERE p.venta_numero=v.venta_numero AND p.venta_tipo=v.venta_tipo AND p.local_codigo=v.local_codigo) AS formas_pago
            FROM ventas v
            LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut
            LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo
            WHERE {where_sql}
            ORDER BY v.venta_fecha DESC, v.venta_numero DESC
            LIMIT %s
        """, tuple(params + [limit]))
        rows = clean_rows(cur.fetchall())
    return json_safe(rows)


@app.get("/api/boletas/{local_codigo}/{venta_numero}")
def boleta_detalle(local_codigo: str, venta_numero: int):
    local = clean_text(local_codigo, 10) or ""
    with cursor() as (_, cur):
        cur.execute("""
            SELECT v.*, COALESCE(c.cliente_nombre, 'CONSUMIDOR FINAL') AS cliente_nombre,
                   COALESCE(ve.vendedor_nombre, v.vendedor_codigo) AS vendedor_nombre
            FROM ventas v
            LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut
            LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo
            WHERE v.venta_numero=%s AND v.venta_tipo='BO' AND v.local_codigo=%s LIMIT 1
        """, (venta_numero, local))
        sale = clean_row(cur.fetchone() or {})
        if not sale:
            raise HTTPException(404, "Boleta no encontrada")
        sale.pop("venta_docadjunto", None)
        sale.pop("venta_xml", None)
        sale.pop("venta_fotodoc", None)
        cur.execute("""
            SELECT d.producto_codigo, COALESCE(NULLIF(d.venta_descripcion,''),p.producto_descripcion,d.producto_codigo) AS producto_descripcion,
                   d.venta_cantidad, d.venta_precio, d.venta_descuentol, d.venta_lineaneto,
                   d.venta_lineaiva, d.venta_lineaila,
                   COALESCE(d.venta_lineaneto,0)+COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0) AS total_linea
            FROM ventaslevel2 d LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo
            WHERE d.venta_numero=%s AND d.venta_tipo='BO' AND d.local_codigo=%s
            ORDER BY producto_descripcion
        """, (venta_numero, local))
        lines = clean_rows(cur.fetchall())
        cur.execute("""
            SELECT p.venta_pagoitem, p.fpago_codigo, COALESCE(fp.fpago_descripcion,p.fpago_codigo) AS fpago_descripcion,
                   p.venta_pagomonto, p.venta_pagoreal, p.venta_numerodoc
            FROM ventaslevel1 p LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo
            WHERE p.venta_numero=%s AND p.venta_tipo='BO' AND p.local_codigo=%s
            ORDER BY p.venta_pagoitem
        """, (venta_numero, local))
        payments = clean_rows(cur.fetchall())
    return json_safe({"venta": sale, "lineas": lines, "pagos": payments})


@app.get("/api/facturas/catalogos")
def facturas_catalogos():
    with cursor() as (_, cur):
        cur.execute("SELECT vendedor_codigo AS value,vendedor_nombre AS label FROM vendedores WHERE COALESCE(vendedor_estado,'A')<>'B' ORDER BY vendedor_nombre")
        sellers = clean_rows(cur.fetchall())
        cur.execute("SELECT CAST(ruta_id AS CHAR) AS value,ruta_nombre AS label FROM rutas ORDER BY ruta_nombre")
        routes = clean_rows(cur.fetchall())
        cur.execute("SELECT vendedor_codigo AS value,vendedor_nombre AS label FROM vendedores WHERE COALESCE(vendedor_estado,'A')<>'B' ORDER BY vendedor_nombre")
        delivery = clean_rows(cur.fetchall())
    return json_safe({"vendedores": sellers, "rutas": routes, "repartidores": delivery})


@app.get("/api/facturas")
def facturas_listado(
    local_codigo: str | None = Query(None), fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None), folio: str | None = Query(None),
    vendedor_codigo: str | None = Query(None), ruta_id: str | None = Query(None),
    repartidor: str | None = Query(None), estado_sii: str | None = Query(None),
    limit: int = Query(250, ge=1, le=500),
):
    filters = ["v.venta_tipo IN ('FE','FA')"]
    params = []
    for value, clause, size in (
        (local_codigo, "v.local_codigo=%s", 10), (vendedor_codigo, "v.vendedor_codigo=%s", 10),
        (repartidor, "v.venta_repartidor=%s", 10), (estado_sii, "COALESCE(v.venta_estadosii,'')=%s", 3),
    ):
        cleaned = clean_text(value, size)
        if cleaned:
            filters.append(clause); params.append(cleaned)
    if fecha_desde:
        filters.append("v.venta_fecha >= %s"); params.append(parse_date(fecha_desde))
    if fecha_hasta:
        filters.append("v.venta_fecha <= %s"); params.append(parse_date(fecha_hasta))
    number = clean_text(folio, 20)
    if number:
        filters.append("(CAST(v.venta_numero AS CHAR) LIKE %s OR CAST(COALESCE(v.venta_folio,0) AS CHAR) LIKE %s)")
        params.extend([f"{number}%", f"{number}%"])
    route = clean_text(ruta_id, 20)
    if route:
        filters.append("CAST(c.ruta_id AS CHAR)=%s"); params.append(route)
    where_sql = " AND ".join(filters)
    base_from = """FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo LEFT JOIN rutas r ON r.ruta_id=c.ruta_id LEFT JOIN vendedores rep ON rep.vendedor_codigo=v.venta_repartidor"""
    with cursor() as (_, cur):
        cur.execute(f"""
            SELECT COUNT(*) documentos,COALESCE(SUM(v.venta_totalventa),0) total,
                   COALESCE(SUM(v.venta_pagototal),0) pagado,
                   COALESCE(SUM(v.venta_totalventa-COALESCE(v.venta_pagototal,0)),0) saldo
            {base_from} WHERE {where_sql}
        """, tuple(params))
        totals = clean_row(cur.fetchone() or {})
        cur.execute(f"""
            SELECT v.venta_numero,v.venta_tipo,v.local_codigo,v.venta_fecha,v.venta_hora,
                   v.venta_folio,v.venta_foliosii,v.venta_estado,v.venta_estadosii,
                   v.venta_impreso,v.venta_entregado,v.venta_enviodte,v.venta_picking,
                   v.cliente_rut,COALESCE(c.cliente_nombre,'SIN CLIENTE') cliente_nombre,
                   c.ruta_id,COALESCE(r.ruta_nombre,CAST(c.ruta_id AS CHAR)) ruta_nombre,
                   v.vendedor_codigo,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) vendedor_nombre,
                   v.venta_repartidor,COALESCE(rep.vendedor_nombre,v.venta_repartidor) repartidor_nombre,
                   v.venta_totalventa,v.venta_pagototal,
                   (v.venta_totalventa-COALESCE(v.venta_pagototal,0)) saldo,
                   (SELECT COUNT(*) FROM ventaslevel2 d WHERE d.venta_numero=v.venta_numero AND d.venta_tipo=v.venta_tipo AND d.local_codigo=v.local_codigo) lineas
            {base_from} WHERE {where_sql}
            ORDER BY v.venta_fecha DESC,COALESCE(v.venta_folio,v.venta_numero) DESC LIMIT %s
        """, tuple(params + [limit]))
        rows = clean_rows(cur.fetchall())
    return json_safe({"rows": rows, "totals": totals})


@app.get("/api/facturas/{local_codigo}/{venta_tipo}/{venta_numero}")
def factura_detalle(local_codigo: str, venta_tipo: str, venta_numero: int):
    local = clean_text(local_codigo,10) or ""; doc_type=clean_text(venta_tipo,2) or "FE"
    with cursor() as (_,cur):
        cur.execute("""SELECT v.*,COALESCE(c.cliente_nombre,'SIN CLIENTE') cliente_nombre,c.cliente_direccion,c.cliente_giro,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) vendedor_nombre FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo WHERE v.venta_numero=%s AND v.venta_tipo=%s AND v.local_codigo=%s LIMIT 1""",(venta_numero,doc_type,local))
        sale=clean_row(cur.fetchone() or {})
        if not sale: raise HTTPException(404,"Factura no encontrada")
        for key in ("venta_docadjunto","venta_xml","venta_fotodoc","venta_token"): sale.pop(key,None)
        cur.execute("""SELECT d.producto_codigo,COALESCE(NULLIF(d.venta_descripcion,''),p.producto_descripcion,d.producto_codigo) producto_descripcion,d.venta_cantidad,d.venta_precio,d.venta_descuentol,d.venta_lineaneto,d.venta_lineaiva,d.venta_lineaila,COALESCE(d.venta_lineaneto,0)+COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0) total_linea FROM ventaslevel2 d LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo WHERE d.venta_numero=%s AND d.venta_tipo=%s AND d.local_codigo=%s ORDER BY producto_descripcion""",(venta_numero,doc_type,local))
        lines=clean_rows(cur.fetchall())
        cur.execute("""SELECT p.venta_pagoitem,p.fpago_codigo,COALESCE(fp.fpago_descripcion,p.fpago_codigo) fpago_descripcion,p.venta_pagomonto,p.venta_numerodoc FROM ventaslevel1 p LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo WHERE p.venta_numero=%s AND p.venta_tipo=%s AND p.local_codigo=%s ORDER BY p.venta_pagoitem""",(venta_numero,doc_type,local))
        payments=clean_rows(cur.fetchall())
    return json_safe({"venta":sale,"lineas":lines,"pagos":payments})


@app.put("/api/facturas/cambio-fecha")
def facturas_cambio_fecha(payload: InvoiceBulkRequest):
    new_date = parse_date(payload.value or "")
    with cursor() as (_,cur):
        for doc in payload.documents:
            cur.execute("UPDATE ventas SET venta_fecha=%s,venta_ult=DAY(%s),venta_mes=MONTH(%s),venta_ano=YEAR(%s) WHERE venta_numero=%s AND venta_tipo=%s AND local_codigo=%s",(new_date,new_date,new_date,new_date,to_int(doc.get("venta_numero"),0),clean_text(doc.get("venta_tipo"),2),clean_text(doc.get("local_codigo"),10)))
    return json_safe({"ok":True,"updated":len(payload.documents)})


@app.put("/api/facturas/asignar-reparto")
def facturas_asignar_reparto(payload: InvoiceBulkRequest):
    delivery=clean_text(payload.value,10) or ""
    with cursor() as (_,cur):
        for doc in payload.documents:
            cur.execute("UPDATE ventas SET venta_repartidor=%s WHERE venta_numero=%s AND venta_tipo=%s AND local_codigo=%s",(delivery,to_int(doc.get("venta_numero"),0),clean_text(doc.get("venta_tipo"),2),clean_text(doc.get("local_codigo"),10)))
    return json_safe({"ok":True,"updated":len(payload.documents)})


@app.put("/api/facturas/enviar-folios")
def facturas_enviar_folios(payload: InvoiceBulkRequest):
    today=datetime.date.today()
    with cursor() as (_,cur):
        for doc in payload.documents:
            cur.execute("UPDATE ventas SET venta_enviodte=%s WHERE venta_numero=%s AND venta_tipo=%s AND local_codigo=%s",(today,to_int(doc.get("venta_numero"),0),clean_text(doc.get("venta_tipo"),2),clean_text(doc.get("local_codigo"),10)))
    return json_safe({"ok":True,"updated":len(payload.documents)})


@app.get("/api/notas-venta/catalogos")
def notas_venta_catalogos():
    with cursor() as (_, cur):
        cur.execute("SELECT vendedor_codigo AS value,vendedor_nombre AS label FROM vendedores WHERE COALESCE(vendedor_estado,'A')<>'B' ORDER BY vendedor_nombre")
        sellers = clean_rows(cur.fetchall())
        cur.execute("SELECT CAST(ruta_id AS CHAR) AS value,ruta_nombre AS label FROM rutas ORDER BY ruta_nombre")
        routes = clean_rows(cur.fetchall())
    return json_safe({"vendedores": sellers, "rutas": routes})


@app.get("/api/notas-venta")
def notas_venta_listado(
    local_codigo: str | None = Query(None), fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None), entrega_desde: str | None = Query(None),
    entrega_hasta: str | None = Query(None), numero: str | None = Query(None),
    vendedor_codigo: str | None = Query(None), ruta_id: str | None = Query(None),
    estado: str | None = Query(None), limit: int = Query(300, ge=1, le=500),
):
    filters = ["v.venta_tipo='NV'"]
    params = []
    for value, clause, size in (
        (local_codigo, "v.local_codigo=%s", 10), (vendedor_codigo, "v.vendedor_codigo=%s", 10),
        (estado, "COALESCE(v.venta_facturado,'N')=%s", 1),
    ):
        cleaned = clean_text(value, size)
        if cleaned:
            filters.append(clause); params.append(cleaned)
    for value, clause in (
        (fecha_desde, "v.venta_fecha>=%s"), (fecha_hasta, "v.venta_fecha<=%s"),
        (entrega_desde, "v.venta_fechavto>=%s"), (entrega_hasta, "v.venta_fechavto<=%s"),
    ):
        if value:
            filters.append(clause); params.append(parse_date(value))
    number = clean_text(numero, 24)
    if number:
        filters.append("CAST(v.venta_numero AS CHAR) LIKE %s"); params.append(f"{number}%")
    route = clean_text(ruta_id, 20)
    if route:
        filters.append("CAST(c.ruta_id AS CHAR)=%s"); params.append(route)
    where_sql = " AND ".join(filters)
    base = "FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo LEFT JOIN rutas r ON r.ruta_id=c.ruta_id"
    with cursor() as (_, cur):
        cur.execute(f"""SELECT COUNT(*) documentos,COALESCE(SUM(CASE WHEN COALESCE(v.venta_facturado,'N')='S' THEN 1 ELSE 0 END),0) facturadas,COALESCE(SUM(CASE WHEN COALESCE(v.venta_facturado,'N')<>'S' THEN 1 ELSE 0 END),0) pendientes,COALESCE(SUM(v.venta_totalventa),0) total {base} WHERE {where_sql}""", tuple(params))
        totals = clean_row(cur.fetchone() or {})
        cur.execute(f"""SELECT v.venta_numero,v.local_codigo,v.venta_fecha,v.venta_hora,v.venta_fechavto,v.venta_tipoemision,v.venta_estado,v.venta_facturado,v.venta_entregado,v.venta_picking,v.cliente_rut,COALESCE(c.cliente_nombre,'SIN CLIENTE') cliente_nombre,c.ruta_id,COALESCE(r.ruta_nombre,CAST(c.ruta_id AS CHAR)) ruta_nombre,v.vendedor_codigo,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) vendedor_nombre,v.venta_totalventa,(SELECT COUNT(*) FROM ventaslevel2 d WHERE d.venta_numero=v.venta_numero AND d.venta_tipo='NV' AND d.local_codigo=v.local_codigo) lineas {base} WHERE {where_sql} ORDER BY v.venta_fecha DESC,v.venta_hora DESC,v.venta_numero DESC LIMIT %s""", tuple(params + [limit]))
        rows = clean_rows(cur.fetchall())
    return json_safe({"rows": rows, "totals": totals})


@app.get("/api/notas-venta/{local_codigo}/{venta_numero}")
def nota_venta_detalle(local_codigo: str, venta_numero: int):
    local = clean_text(local_codigo, 10) or ""
    with cursor() as (_, cur):
        cur.execute("""SELECT v.*,COALESCE(c.cliente_nombre,'SIN CLIENTE') cliente_nombre,c.cliente_direccion,c.cliente_giro,c.ruta_id,COALESCE(r.ruta_nombre,CAST(c.ruta_id AS CHAR)) ruta_nombre,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) vendedor_nombre FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN rutas r ON r.ruta_id=c.ruta_id LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo WHERE v.venta_numero=%s AND v.venta_tipo='NV' AND v.local_codigo=%s LIMIT 1""", (venta_numero, local))
        sale = clean_row(cur.fetchone() or {})
        if not sale:
            raise HTTPException(404, "Nota de venta no encontrada")
        for key in ("venta_docadjunto", "venta_xml", "venta_fotodoc", "venta_token"):
            sale.pop(key, None)
        cur.execute("""SELECT d.producto_codigo,COALESCE(NULLIF(d.venta_descripcion,''),p.producto_descripcion,d.producto_codigo) producto_descripcion,d.venta_cantidad,d.venta_descuentol,d.venta_precio,d.venta_lineaneto,d.venta_lineaiva,d.venta_lineaila,COALESCE(d.venta_lineaneto,0)+COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0) total_linea FROM ventaslevel2 d LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo WHERE d.venta_numero=%s AND d.venta_tipo='NV' AND d.local_codigo=%s ORDER BY producto_descripcion""", (venta_numero, local))
        lines = clean_rows(cur.fetchall())
    return json_safe({"venta": sale, "lineas": lines})


def save_nota_venta(cur, payload: NvSaveRequest, venta_numero: int | None = None):
    local = clean_text(payload.local_codigo, 10) or ""
    client = clean_text(payload.cliente_rut, 20) or ""
    seller = clean_text(payload.vendedor_codigo, 10) or ""
    issued = parse_date(payload.fecha_emision)
    delivery = parse_date(payload.fecha_entrega)
    if delivery < issued:
        raise HTTPException(400, "La fecha de entrega no puede ser anterior a la emision")
    if not payload.lines:
        raise HTTPException(400, "Agregue al menos un producto")
    cur.execute("SELECT cliente_direccion,cliente_condiccion FROM clientes WHERE cliente_rut=%s LIMIT 1", (client,))
    client_row = clean_row(cur.fetchone() or {})
    if not client_row:
        raise HTTPException(400, "Cliente no encontrado")
    details = []
    total_neto = total_iva = total_ila = total_sale = 0
    for requested in payload.lines:
        product = pos_product(cur, requested.producto_codigo, local)
        if not product:
            raise HTTPException(400, f"Producto {requested.producto_codigo} no encontrado")
        quantity = decimal_value(requested.cantidad)
        discount = max(decimal.Decimal("0"), min(decimal.Decimal("100"), decimal_value(requested.descuento)))
        if quantity <= 0:
            raise HTTPException(400, "La cantidad debe ser mayor que cero")
        factor = decimal.Decimal("1") - discount / decimal.Decimal("100")
        neto = round_business(decimal_value(product.get("precio_neto")) * quantity * factor)
        iva = round_business(decimal_value(product.get("precio_iva")) * quantity * factor)
        ila = round_business(decimal_value(product.get("precio_ila")) * quantity * factor)
        total = round_business(decimal_value(product.get("precio_venta")) * quantity * factor)
        total_neto += neto; total_iva += iva; total_ila += ila; total_sale += total
        details.append((product, quantity, discount, neto, iva, ila, total))
    if venta_numero is None:
        cur.execute("""CREATE TABLE IF NOT EXISTS app_document_sequences(document_type VARCHAR(10) NOT NULL,current_number BIGINT NOT NULL,PRIMARY KEY(document_type)) ENGINE=InnoDB DEFAULT CHARSET=latin1""")
        cur.execute("INSERT IGNORE INTO app_document_sequences(document_type,current_number) VALUES('NV',0)")
        cur.execute("SELECT current_number FROM app_document_sequences WHERE document_type='NV' FOR UPDATE")
        current = to_int((cur.fetchone() or {}).get("current_number"), 0)
        if current <= 0:
            cur.execute("SELECT venta_numero FROM ventas WHERE venta_tipo='NV' ORDER BY venta_numero DESC LIMIT 1")
            current = max(311225000000, to_int((cur.fetchone() or {}).get("venta_numero"), 0))
        venta_numero = current + 1
        cur.execute("UPDATE app_document_sequences SET current_number=%s WHERE document_type='NV'", (venta_numero,))
        cur.execute("""INSERT INTO ventas(venta_numero,venta_tipo,local_codigo,venta_fecha,cliente_rut,venta_ult,venta_mes,venta_ano,venta_estado,vendedor_codigo,venta_descuento,venta_totalventa,venta_observacion01,venta_hora,venta_fechavto,venta_condicion,venta_direccion,venta_usuario,venta_neto1,venta_iva1,venta_ila1,venta_facturado,venta_entregado,venta_tipoemision) VALUES(%s,'NV',%s,%s,%s,DAY(%s),MONTH(%s),YEAR(%s),'A',%s,0,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'N','N','F')""", (venta_numero, local, issued, client, issued, issued, issued, seller, total_sale, clean_text(payload.observacion, 250) or "", datetime.datetime.now(), delivery, clean_text(client_row.get("cliente_condiccion"), 10) or "", clean_text(client_row.get("cliente_direccion"), 100) or "", payload.user_id, total_neto, total_iva, total_ila))
    else:
        cur.execute("SELECT venta_facturado FROM ventas WHERE venta_numero=%s AND venta_tipo='NV' AND local_codigo=%s FOR UPDATE", (venta_numero, local))
        existing = clean_row(cur.fetchone() or {})
        if not existing:
            raise HTTPException(404, "Nota de venta no encontrada")
        if clean_text(existing.get("venta_facturado"), 1) == "S":
            raise HTTPException(409, "Una nota de venta facturada no puede modificarse")
        cur.execute("""UPDATE ventas SET venta_fecha=%s,cliente_rut=%s,venta_ult=DAY(%s),venta_mes=MONTH(%s),venta_ano=YEAR(%s),vendedor_codigo=%s,venta_totalventa=%s,venta_observacion01=%s,venta_fechavto=%s,venta_condicion=%s,venta_direccion=%s,venta_usuario=%s,venta_neto1=%s,venta_iva1=%s,venta_ila1=%s WHERE venta_numero=%s AND venta_tipo='NV' AND local_codigo=%s""", (issued, client, issued, issued, issued, seller, total_sale, clean_text(payload.observacion, 250) or "", delivery, clean_text(client_row.get("cliente_condiccion"), 10) or "", clean_text(client_row.get("cliente_direccion"), 100) or "", payload.user_id, total_neto, total_iva, total_ila, venta_numero, local))
        cur.execute("DELETE FROM ventaslevel2 WHERE venta_numero=%s AND venta_tipo='NV' AND local_codigo=%s", (venta_numero, local))
    cur.execute("SELECT local_bodega FROM locales WHERE local_codigo=%s LIMIT 1", (local,))
    bodega = clean_text((cur.fetchone() or {}).get("local_bodega"), 10) or local
    for product, quantity, discount, neto, iva, ila, total in details:
        code = clean_text(product.get("producto_codigo"), 20) or ""
        cur.execute("""INSERT INTO ventaslevel2(venta_numero,venta_tipo,local_codigo,producto_codigo,bodega_codigo,venta_lineaila,venta_lineaiva,venta_lineaneto,venta_precio,venta_cantidad,empleado_codigo,venta_descuentol,venta_unidadenvase,venta_descripcion,venta_precioventa,venta_totalneto,venta_cajas) VALUES(%s,'NV',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0)""", (venta_numero, local, code, bodega, ila, iva, neto, to_int(product.get("precio_neto"), 0), float(quantity), seller, float(discount), to_float(product.get("producto_unidadenvase"), 0), clean_text(product.get("producto_descripcion"), 50) or code, to_int(product.get("precio_venta"), 0), neto))
    return venta_numero, total_sale


@app.post("/api/notas-venta")
def nota_venta_crear(payload: NvSaveRequest):
    with cursor() as (_, cur):
        number, total = save_nota_venta(cur, payload)
    return json_safe({"ok": True, "venta_numero": number, "total": total})


@app.put("/api/notas-venta/{venta_numero}")
def nota_venta_actualizar(venta_numero: int, payload: NvSaveRequest):
    with cursor() as (_, cur):
        number, total = save_nota_venta(cur, payload, venta_numero)
    return json_safe({"ok": True, "venta_numero": number, "total": total})


@app.delete("/api/notas-venta/{local_codigo}/{venta_numero}")
def nota_venta_anular(local_codigo: str, venta_numero: int):
    local = clean_text(local_codigo, 10) or ""
    with cursor() as (_, cur):
        cur.execute("UPDATE ventas SET venta_estado='N' WHERE venta_numero=%s AND venta_tipo='NV' AND local_codigo=%s AND COALESCE(venta_facturado,'N')<>'S'", (venta_numero, local))
        if cur.rowcount == 0:
            raise HTTPException(409, "La nota no existe o ya fue facturada")
    return json_safe({"ok": True})


@app.post("/api/notas-venta/{local_codigo}/{venta_numero}/copiar")
def nota_venta_copiar(local_codigo: str, venta_numero: int, request: dict):
    local = clean_text(local_codigo, 10) or ""
    with cursor() as (_, cur):
        cur.execute("SELECT cliente_rut,vendedor_codigo,venta_fechavto,venta_observacion01,venta_fecha FROM ventas WHERE venta_numero=%s AND venta_tipo='NV' AND local_codigo=%s", (venta_numero, local))
        source = clean_row(cur.fetchone() or {})
        if not source:
            raise HTTPException(404, "Nota de venta no encontrada")
        cur.execute("SELECT producto_codigo,venta_cantidad,venta_descuentol FROM ventaslevel2 WHERE venta_numero=%s AND venta_tipo='NV' AND local_codigo=%s", (venta_numero, local))
        lines = clean_rows(cur.fetchall())
        payload = NvSaveRequest(local_codigo=local, user_id=to_int(request.get("user_id"), 0), cliente_rut=source.get("cliente_rut") or "", vendedor_codigo=source.get("vendedor_codigo") or "", fecha_emision=str(datetime.date.today()), fecha_entrega=str(max(datetime.date.today(), source.get("venta_fechavto") or datetime.date.today())), observacion=f"Copia NV {venta_numero}", lines=[NvLineRequest(producto_codigo=x.get("producto_codigo") or "", cantidad=to_float(x.get("venta_cantidad"), 0), descuento=to_float(x.get("venta_descuentol"), 0)) for x in lines])
        number, total = save_nota_venta(cur, payload)
    return json_safe({"ok": True, "venta_numero": number, "total": total})


def management_data(cur, year: int, month: int, local_codigo: str | None = None):
    start = datetime.date(year, month, 1)
    end = datetime.date(year + (month == 12), 1 if month == 12 else month + 1, 1)
    local = clean_text(local_codigo, 10)
    local_filter = " AND v.local_codigo=%s" if local else ""
    params = [start, end] + ([local] if local else [])
    signed_total = "CASE WHEN v.venta_tipo='NC' THEN -COALESCE(v.venta_totalventa,0) ELSE COALESCE(v.venta_totalventa,0) END"
    signed_neto = "CASE WHEN v.venta_tipo='NC' THEN -COALESCE(v.venta_neto1,0) ELSE COALESCE(v.venta_neto1,0) END"
    signed_paid = "CASE WHEN v.venta_tipo='NC' THEN -COALESCE(v.venta_pagototal,0) ELSE COALESCE(v.venta_pagototal,0) END"
    base_where = f"v.venta_fecha>=%s AND v.venta_fecha<%s AND v.venta_tipo IN ('FE','FA','BO','NC'){local_filter}"
    cur.execute(f"SELECT COUNT(*) documentos,SUM({signed_neto}) neto,SUM({signed_total}) total,SUM({signed_paid}) pagado FROM ventas v WHERE {base_where}", tuple(params))
    summary = clean_row(cur.fetchone() or {})
    summary["saldo"] = to_float(summary.get("total"),0)-to_float(summary.get("pagado"),0)
    cur.execute(f"SELECT DATE_FORMAT(v.venta_fecha,'%Y-%m-%d') label,SUM({signed_neto}) neto,SUM({signed_total}) total,SUM({signed_paid}) pagado FROM ventas v WHERE {base_where} GROUP BY v.venta_fecha ORDER BY v.venta_fecha", tuple(params)); daily=clean_rows(cur.fetchall())
    cur.execute(f"SELECT v.local_codigo codigo,COALESCE(l.local_descripcion,v.local_codigo) label,SUM({signed_neto}) neto,SUM({signed_total}) total,SUM({signed_paid}) pagado FROM ventas v LEFT JOIN locales l ON l.local_codigo=v.local_codigo WHERE {base_where} GROUP BY v.local_codigo,l.local_descripcion ORDER BY total DESC", tuple(params)); locals_data=clean_rows(cur.fetchall())
    cur.execute(f"SELECT v.venta_tipo codigo,v.venta_tipo label,SUM({signed_neto}) neto,SUM({signed_total}) total,SUM({signed_paid}) pagado FROM ventas v WHERE {base_where} GROUP BY v.venta_tipo ORDER BY total DESC", tuple(params)); documents=clean_rows(cur.fetchall())
    cur.execute(f"SELECT v.vendedor_codigo codigo,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) label,COUNT(*) documentos,SUM({signed_neto}) neto,SUM({signed_total}) total,SUM({signed_paid}) pagado FROM ventas v LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo WHERE {base_where} GROUP BY v.vendedor_codigo,ve.vendedor_nombre ORDER BY total DESC LIMIT 20", tuple(params)); sellers=clean_rows(cur.fetchall())
    cur.execute(f"SELECT COALESCE(r.ruta_nombre,'SIN RUTA') label,COUNT(*) documentos,SUM({signed_total}) total,SUM({signed_paid}) pagado FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN rutas r ON r.ruta_id=c.ruta_id WHERE {base_where} GROUP BY c.ruta_id,r.ruta_nombre ORDER BY total DESC LIMIT 12", tuple(params)); routes=clean_rows(cur.fetchall())
    cur.execute(f"SELECT COALESCE(f.familia_descripcion,'SIN FAMILIA') label,SUM(COALESCE(d.venta_lineaneto,0)) neto,SUM(COALESCE(d.venta_lineaneto,0)+COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0)) total FROM ventaslevel2 d INNER JOIN ventas v ON v.venta_numero=d.venta_numero AND v.venta_tipo=d.venta_tipo AND v.local_codigo=d.local_codigo LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo LEFT JOIN familias f ON f.familia_codigo=p.familia_codigo WHERE {base_where} GROUP BY p.familia_codigo,f.familia_descripcion ORDER BY total DESC LIMIT 12", tuple(params)); families=clean_rows(cur.fetchall())
    cur.execute(f"SELECT COALESCE(fp.fpago_descripcion,p.fpago_codigo) label,COUNT(*) operaciones,SUM(COALESCE(p.venta_pagomonto,0)) total FROM ventaslevel1 p INNER JOIN ventas v ON v.venta_numero=p.venta_numero AND v.venta_tipo=p.venta_tipo AND v.local_codigo=p.local_codigo LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo WHERE {base_where} GROUP BY p.fpago_codigo,fp.fpago_descripcion ORDER BY total DESC", tuple(params)); payments=clean_rows(cur.fetchall())
    compare=[]; compare_filter=" AND local_codigo=%s" if local else ""
    for selected_year in (year-1,year):
        values={}
        for selected_month in range(1,13):
            point_start=datetime.date(selected_year,selected_month,1); point_end=datetime.date(selected_year+(selected_month==12),1 if selected_month==12 else selected_month+1,1)
            cur.execute(f"SELECT SUM(CASE WHEN venta_tipo='NC' THEN -COALESCE(venta_totalventa,0) ELSE COALESCE(venta_totalventa,0) END) total FROM ventas FORCE INDEX (Uventas1) WHERE venta_fecha>=%s AND venta_fecha<%s AND venta_tipo IN ('FE','FA','BO','NC'){compare_filter}",tuple([point_start,point_end]+([local] if local else [])))
            values[str(selected_month)]=to_int((cur.fetchone() or {}).get('total'),0)
        compare.append({"year":selected_year,"values":values})
    return {"year":year,"month":month,"local_codigo":local or "","summary":summary,"daily":daily,"locals":locals_data,"documents":documents,"sellers":sellers,"routes":routes,"families":families,"payments":payments,"comparison":compare}


MANAGEMENT_CACHE = {}


def cached_management_data(year: int, month: int, local_codigo: str | None = None):
    key=(year,month,clean_text(local_codigo,10) or ""); now=datetime.datetime.now()
    cached=MANAGEMENT_CACHE.get(key)
    if cached and (now-cached[0]).total_seconds()<300: return cached[1]
    with cursor() as (_,cur): data=management_data(cur,year,month,local_codigo)
    MANAGEMENT_CACHE[key]=(now,data)
    return data


@app.get("/api/gerencia/dashboard")
def gerencia_dashboard(year: int = Query(..., ge=2000, le=2100), month: int = Query(..., ge=1, le=12), local_codigo: str | None = Query(None)):
    return json_safe(cached_management_data(year,month,local_codigo))


@app.get("/api/gerencia/export/{format_name}")
def gerencia_export(format_name: str, year: int = Query(..., ge=2000, le=2100), month: int = Query(..., ge=1, le=12), local_codigo: str | None = Query(None)):
    data=cached_management_data(year,month,local_codigo)
    title=f"RISEK Gerencia {year}-{month:02d}"
    if format_name == "excel":
        wb=Workbook(); ws=wb.active; ws.title="Dashboard"
        ws.append([title]); ws.merge_cells("A1:F1"); ws["A1"].font=Font(size=18,bold=True,color="FFFFFF"); ws["A1"].fill=PatternFill("solid",fgColor="174A5B"); ws["A1"].alignment=Alignment(horizontal="center")
        ws.append([]); ws.append(["Indicador","Neto","Total","Pagado","Saldo","Documentos"]); s=data['summary']; ws.append(["Periodo",s.get('neto',0),s.get('total',0),s.get('pagado',0),s.get('saldo',0),s.get('documentos',0)])
        for cell in ws[3]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2D7285")
        for col in "BCDE": ws[f"{col}4"].number_format='#,##0'
        ws.append([]); ws.append(["Vendedor","Documentos","Neto","Total","Pagado"])
        seller_start=ws.max_row+1
        for row in data['sellers']: ws.append([row.get('label'),row.get('documentos'),row.get('neto'),row.get('total'),row.get('pagado')])
        for cell in ws[seller_start-1]: cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2D7285")
        for row in range(seller_start,ws.max_row+1):
            for col in range(3,6): ws.cell(row,col).number_format='#,##0'
        chart=BarChart(); chart.title="Ventas mensuales por vendedor"; chart.y_axis.title="Total"; chart.add_data(Reference(ws,min_col=4,min_row=seller_start-1,max_row=ws.max_row),titles_from_data=True); chart.set_categories(Reference(ws,min_col=1,min_row=seller_start,max_row=ws.max_row)); chart.height=8; chart.width=15; ws.add_chart(chart,"H3")
        detail=wb.create_sheet("Ventas diarias"); detail.append(["Fecha","Neto","Total","Pagado"])
        for row in data['daily']: detail.append([row.get('label'),row.get('neto'),row.get('total'),row.get('pagado')])
        line=LineChart(); line.title="Evolucion diaria"; line.add_data(Reference(detail,min_col=2,min_row=1,max_col=4,max_row=detail.max_row),titles_from_data=True); line.set_categories(Reference(detail,min_col=1,min_row=2,max_row=detail.max_row)); line.height=8; line.width=16; detail.add_chart(line,"F2")
        for sheet in wb.worksheets:
            sheet.freeze_panes="A2"; sheet.sheet_view.showGridLines=False
            for index,column in enumerate(sheet.columns,start=1):
                letter=get_column_letter(index); sheet.column_dimensions[letter].width=min(30,max(12,max(len(str(c.value or '')) for c in column)+2))
        output=io.BytesIO(); wb.save(output)
        return Response(content=output.getvalue(),media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="gerencia-{year}-{month:02d}.xlsx"'})
    if format_name == "pdf":
        output=io.BytesIO(); doc=SimpleDocTemplate(output,pagesize=landscape(A4),rightMargin=12*mm,leftMargin=12*mm,topMargin=10*mm,bottomMargin=10*mm); styles=getSampleStyleSheet(); story=[Paragraph(title,styles['Title']),Spacer(1,5*mm)]
        s=data['summary']; summary_table=Table([["NETO","TOTAL","PAGADO","SALDO","DOCUMENTOS"],[f"${to_int(s.get('neto'),0):,}",f"${to_int(s.get('total'),0):,}",f"${to_int(s.get('pagado'),0):,}",f"${to_int(s.get('saldo'),0):,}",str(to_int(s.get('documentos'),0))]],colWidths=[48*mm]*5); summary_table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#174A5B')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('ALIGN',(0,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),.5,colors.HexColor('#CBD9DE')),('PADDING',(0,0),(-1,-1),8)])); story+=[summary_table,Spacer(1,6*mm),Paragraph("Ventas por vendedor",styles['Heading2'])]
        rows=[["Vendedor","Documentos","Neto","Total","Pagado"]]+[[x.get('label'),str(x.get('documentos')),f"${to_int(x.get('neto'),0):,}",f"${to_int(x.get('total'),0):,}",f"${to_int(x.get('pagado'),0):,}"] for x in data['sellers']]
        table=Table(rows,colWidths=[70*mm,28*mm,42*mm,42*mm,42*mm],repeatRows=1); table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2D7285')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('ALIGN',(1,1),(-1,-1),'RIGHT'),('GRID',(0,0),(-1,-1),.35,colors.HexColor('#DCE6E9')),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F3F7F8')]) ])); story.append(table); doc.build(story)
        return Response(content=output.getvalue(),media_type="application/pdf",headers={"Content-Disposition":f'attachment; filename="gerencia-{year}-{month:02d}.pdf"'})
    raise HTTPException(404,"Formato no disponible")


@app.get("/api/compras")
def compras_listado(fecha_desde: str | None=Query(None),fecha_hasta: str | None=Query(None),numero: str | None=Query(None),proveedor_codigo: str | None=Query(None),local_codigo: str | None=Query(None),estado: str | None=Query(None),limit:int=Query(300,ge=1,le=500)):
    filters=[]; params=[]
    for value,clause,size in ((proveedor_codigo,"c.proveedor_codigo=%s",20),(local_codigo,"c.local_codigo=%s",10),(estado,"c.compra_estado=%s",1)):
        cleaned=clean_text(value,size)
        if cleaned: filters.append(clause); params.append(cleaned)
    if fecha_desde: filters.append("c.compra_fecha>=%s"); params.append(parse_date(fecha_desde))
    if fecha_hasta: filters.append("c.compra_fecha<=%s"); params.append(parse_date(fecha_hasta))
    if numero: filters.append("CAST(c.compra_numero AS CHAR) LIKE %s"); params.append(f"{clean_text(numero,24)}%")
    where=" AND ".join(filters) or "1=1"
    with cursor() as (_,cur):
        ensure_purchase_receipts(cur)
        cur.execute(f"""SELECT c.compra_numero,c.compra_tipo,c.local_codigo,COALESCE(l.local_descripcion,c.local_codigo) local_nombre,c.compra_fecha,c.proveedor_codigo,COALESCE(p.proveedor_nombre,c.proveedor_codigo) proveedor_nombre,c.compra_neto,c.compra_iva,c.compra_ila,c.compra_totalcompra,c.compra_estado,c.compra_estadopago,COALESCE((SELECT SUM(pc.compra_pagomonto) FROM pagocompraslevel2 pc WHERE pc.compra_numero=c.compra_numero AND pc.compra_tipo=c.compra_tipo AND pc.local_codigo=c.local_codigo AND pc.proveedor_codigo=c.proveedor_codigo),0) pagado,(SELECT COUNT(*) FROM compraslevel1 d WHERE d.compra_numero=c.compra_numero AND d.compra_tipo=c.compra_tipo AND d.local_codigo=c.local_codigo AND d.proveedor_codigo=c.proveedor_codigo) lineas,EXISTS(SELECT 1 FROM app_purchase_receipts ar WHERE ar.compra_numero=c.compra_numero AND ar.compra_tipo=c.compra_tipo AND ar.local_codigo=c.local_codigo AND ar.proveedor_codigo=c.proveedor_codigo) inventariado FROM compras c LEFT JOIN proveedores p ON p.proveedor_codigo=c.proveedor_codigo LEFT JOIN locales l ON l.local_codigo=c.local_codigo WHERE {where} ORDER BY c.compra_fecha DESC,c.compra_numero DESC LIMIT %s""",tuple(params+[limit])); rows=clean_rows(cur.fetchall())
    return json_safe(rows)


@app.get("/api/compras/catalogos")
def compras_catalogos():
    with cursor() as (_,cur):
        cur.execute("SELECT proveedor_codigo value,proveedor_nombre label FROM proveedores ORDER BY proveedor_nombre"); providers=clean_rows(cur.fetchall())
    return json_safe({"proveedores":providers})


@app.get("/api/compras/{local_codigo}/{compra_tipo}/{proveedor_codigo}/{compra_numero}")
def compra_detalle(local_codigo:str,compra_tipo:str,proveedor_codigo:str,compra_numero:int):
    keys=(compra_numero,clean_text(compra_tipo,2),clean_text(local_codigo,10),clean_text(proveedor_codigo,20))
    with cursor() as (_,cur):
        cur.execute("""SELECT c.*,COALESCE(p.proveedor_nombre,c.proveedor_codigo) proveedor_nombre,COALESCE(l.local_descripcion,c.local_codigo) local_nombre FROM compras c LEFT JOIN proveedores p ON p.proveedor_codigo=c.proveedor_codigo LEFT JOIN locales l ON l.local_codigo=c.local_codigo WHERE c.compra_numero=%s AND c.compra_tipo=%s AND c.local_codigo=%s AND c.proveedor_codigo=%s""",keys); purchase=clean_row(cur.fetchone() or {})
        if not purchase: raise HTTPException(404,"Compra no encontrada")
        cur.execute("""SELECT d.*,COALESCE(p.producto_descripcion,NULLIF(d.compra_descripcion,''),d.producto_codigo) producto_descripcion,COALESCE(b.bodega_descripcion,d.bodega_codigo) bodega_nombre,ROUND(COALESCE(d.compra_cantidadt,d.compra_cantidad,0)*COALESCE(d.compra_valor,0),0) total_linea FROM compraslevel1 d LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo LEFT JOIN bodegas b ON b.bodega_codigo=d.bodega_codigo WHERE d.compra_numero=%s AND d.compra_tipo=%s AND d.local_codigo=%s AND d.proveedor_codigo=%s ORDER BY producto_descripcion""",keys); lines=clean_rows(cur.fetchall())
    return json_safe({"compra":purchase,"lineas":lines})


def ensure_purchase_receipts(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS app_purchase_receipts(compra_numero BIGINT NOT NULL,compra_tipo CHAR(2) NOT NULL,local_codigo CHAR(10) NOT NULL,proveedor_codigo CHAR(20) NOT NULL,closed_at DATETIME NOT NULL,user_id INT NOT NULL DEFAULT 0,line_count INT NOT NULL DEFAULT 0,total_quantity DECIMAL(16,3) NOT NULL DEFAULT 0,PRIMARY KEY(compra_numero,compra_tipo,local_codigo,proveedor_codigo)) ENGINE=InnoDB DEFAULT CHARSET=latin1""")


def close_purchase_inventory(cur, local, doc_type, provider, compra_numero, user_id=0):
    keys = (compra_numero, doc_type, local, provider)
    now = datetime.datetime.now()
    ensure_purchase_receipts(cur)
    cur.execute(
        "SELECT compra_fecha,compra_estado FROM compras WHERE compra_numero=%s AND compra_tipo=%s AND local_codigo=%s AND proveedor_codigo=%s FOR UPDATE",
        keys,
    )
    purchase = cur.fetchone() or {}
    if not purchase:
        raise HTTPException(404, "Compra no encontrada")
    cur.execute(
        "SELECT 1 FROM app_purchase_receipts WHERE compra_numero=%s AND compra_tipo=%s AND local_codigo=%s AND proveedor_codigo=%s",
        keys,
    )
    if cur.fetchone():
        return {"ok": True, "already_closed": True}
    if (clean_text(purchase.get("compra_estado"), 1) or "") != "A":
        raise HTTPException(409, "Solo las compras abiertas pueden cerrarse e ingresar a inventario")

    cur.execute(
        """SELECT producto_codigo,bodega_codigo,
                  COALESCE(compra_cantidadt,compra_cantidad,0) cantidad,
                  compra_valor,compra_pespecial,compra_codigoprov
           FROM compraslevel1
           WHERE compra_numero=%s AND compra_tipo=%s AND local_codigo=%s AND proveedor_codigo=%s
           FOR UPDATE""",
        keys,
    )
    lines = cur.fetchall()
    if not lines:
        raise HTTPException(400, "La compra no contiene productos")

    purchase_date = purchase.get("compra_fecha") or datetime.date.today()
    if isinstance(purchase_date, datetime.datetime):
        purchase_date = purchase_date.date()
    elif not isinstance(purchase_date, datetime.date):
        purchase_date = parse_date(str(purchase_date))
    month, year = purchase_date.month, purchase_date.year
    price_list = "30" if local == "02" else "01"
    total_quantity = decimal.Decimal("0")
    processed = 0

    for line in lines:
        product = clean_text(line.get("producto_codigo"), 20) or ""
        warehouse = clean_text(line.get("bodega_codigo"), 10) or local
        quantity = decimal_value(line.get("cantidad"))
        if not product or quantity <= 0:
            continue

        cur.execute("SELECT * FROM productos WHERE producto_codigo=%s FOR UPDATE", (product,))
        current = cur.fetchone() or {}
        if not current:
            raise HTTPException(409, f"El producto {product} no existe")

        cur.execute(
            "SELECT COALESCE(producto_stockbodega,0) stock FROM productoslevel2 WHERE producto_codigo=%s AND bodega_codigo=%s FOR UPDATE",
            (product, warehouse),
        )
        warehouse_row = cur.fetchone() or {}
        previous_stock = decimal_value(warehouse_row.get("stock"))
        old_cost = decimal_value(current.get("producto_costo"))
        new_cost = decimal_value(line.get("compra_valor"))
        special_cost = decimal_value(line.get("compra_pespecial"))
        supplier_code = clean_text(line.get("compra_codigoprov"), 20) or ""

        positive_stock = max(previous_stock, decimal.Decimal("0"))
        average_units = positive_stock + quantity
        average_cost = new_cost
        if average_units > 0:
            average_cost = ((positive_stock * old_cost) + (quantity * new_cost)) / average_units

        incoming = {
            "producto_costo": new_cost,
            "producto_costoof": special_cost,
        }
        calculated, tax_rate = producto_business_values(cur, product, incoming, current, False)
        product_updates = {
            "producto_costoant": round_business(old_cost),
            "producto_costo": new_cost,
            "producto_costoof": special_cost,
            "producto_cup": round_business(average_cost),
            "producto_proveedor": provider,
            **calculated,
        }
        if supplier_code:
            product_updates["producto_alternativo"] = supplier_code
        set_sql = ", ".join(f"`{column}`=%s" for column in product_updates)
        cur.execute(
            f"UPDATE productos SET {set_sql} WHERE producto_codigo=%s",
            tuple(product_updates.values()) + (product,),
        )

        neto = decimal_value(calculated["producto_neto"])
        price_values = {
            "lista_neto": int(neto),
            "lista_iva": calculated["producto_iva"],
            "lista_ila": calculated["producto_ila"],
            "lista_impuesto": tax_rate,
            "lista_venta": calculated["producto_venta"],
            "lista_precio": calculated["producto_venta"],
            "lista_fecha": purchase_date,
            "Lista_costo": new_cost,
            "lista_margen": float(((neto / new_cost) - 1) * 100) if new_cost else 0,
        }
        price_columns = ["lista_codigo", "producto_codigo"] + list(price_values)
        price_data = [price_list, product] + list(price_values.values())
        price_updates = ", ".join(f"`{column}`=VALUES(`{column}`)" for column in price_values)
        cur.execute(
            f"INSERT INTO precioslevel1 ({', '.join(f'`{column}`' for column in price_columns)}) "
            f"VALUES ({', '.join(['%s'] * len(price_columns))}) ON DUPLICATE KEY UPDATE {price_updates}",
            tuple(price_data),
        )

        cur.execute(
            """INSERT INTO stock_mensuales
               (producto_codigo,bodega_codigo,stock_mes,stock_ano,stock_fecha,stock_hora,stock_inicial)
               VALUES(%s,%s,%s,%s,%s,%s,0)
               ON DUPLICATE KEY UPDATE stock_fecha=VALUES(stock_fecha),stock_hora=VALUES(stock_hora)""",
            (product, warehouse, month, year, purchase_date, now),
        )
        cur.execute(
            """INSERT INTO stock_mensualeslevel1
               (producto_codigo,bodega_codigo,stock_mes,stock_ano,stock_folio,stock_tm,stock_tipo,stock_fechadoc,stock_grabacion,stock_cantidad)
               VALUES(%s,%s,%s,%s,%s,'I',%s,%s,%s,%s)
               ON DUPLICATE KEY UPDATE stock_fechadoc=VALUES(stock_fechadoc),stock_grabacion=VALUES(stock_grabacion),stock_cantidad=VALUES(stock_cantidad)""",
            (product, warehouse, month, year, compra_numero, doc_type, purchase_date, now, quantity),
        )

        cur.execute(
            """SELECT cartola_item FROM cartolalevel1
               WHERE producto_codigo=%s AND bodega_codigo=%s AND cartola_fecha=%s
                 AND cartola_tipodoc=%s AND cartola_numero=%s
               ORDER BY cartola_item LIMIT 1 FOR UPDATE""",
            (product, warehouse, purchase_date, doc_type, compra_numero),
        )
        cartola_row = cur.fetchone()
        if cartola_row:
            cur.execute(
                """UPDATE cartolalevel1 SET cartola_ingreso=%s,cartola_salida=0,cartola_hora=%s
                   WHERE producto_codigo=%s AND bodega_codigo=%s AND cartola_item=%s""",
                (quantity, now, product, warehouse, cartola_row["cartola_item"]),
            )
        else:
            cur.execute(
                "SELECT COALESCE(MAX(cartola_item),0)+1 next_item FROM cartolalevel1 WHERE producto_codigo=%s AND bodega_codigo=%s FOR UPDATE",
                (product, warehouse),
            )
            cartola_item = (cur.fetchone() or {}).get("next_item", 1)
            cur.execute(
                """INSERT INTO cartolalevel1
                   (producto_codigo,bodega_codigo,cartola_item,cartola_ingreso,cartola_salida,cartola_fecha,cartola_tipodoc,cartola_numero,cartola_hora)
                   VALUES(%s,%s,%s,%s,0,%s,%s,%s,%s)""",
                (product, warehouse, cartola_item, quantity, purchase_date, doc_type, compra_numero, now),
            )

        cur.execute(
            """SELECT COALESCE(SUM(COALESCE(cartola_ingreso,0)-COALESCE(cartola_salida,0)),0) saldo
               FROM cartolalevel1 WHERE producto_codigo=%s AND bodega_codigo=%s""",
            (product, warehouse),
        )
        cartola_balance = decimal_value((cur.fetchone() or {}).get("saldo"))
        cur.execute(
            """INSERT INTO productoslevel2(producto_codigo,bodega_codigo,producto_stockbodega,producto_reservado)
               VALUES(%s,%s,%s,0)
               ON DUPLICATE KEY UPDATE producto_stockbodega=VALUES(producto_stockbodega)""",
            (product, warehouse, cartola_balance),
        )
        total_quantity += quantity
        processed += 1

    if not processed:
        raise HTTPException(400, "La compra no contiene cantidades válidas para ingresar")
    cur.execute(
        """UPDATE compraslevel1 SET compra_recepcion=COALESCE(compra_cantidadt,compra_cantidad,0),compra_entrega='S'
           WHERE compra_numero=%s AND compra_tipo=%s AND local_codigo=%s AND proveedor_codigo=%s""",
        keys,
    )
    cur.execute(
        "UPDATE compras SET compra_estado='C',compra_enviado='S' WHERE compra_numero=%s AND compra_tipo=%s AND local_codigo=%s AND proveedor_codigo=%s",
        keys,
    )
    cur.execute(
        """INSERT INTO app_purchase_receipts
           (compra_numero,compra_tipo,local_codigo,proveedor_codigo,closed_at,user_id,line_count,total_quantity)
           VALUES(%s,%s,%s,%s,%s,%s,%s,%s)""",
        (compra_numero, doc_type, local, provider, now, to_int(user_id, 0), processed, total_quantity),
    )
    return {"ok": True, "lineas": processed, "cantidad": clean_value(total_quantity)}


@app.post("/api/compras/{local_codigo}/{compra_tipo}/{proveedor_codigo}/{compra_numero}/cerrar")
def compra_cerrar(local_codigo:str,compra_tipo:str,proveedor_codigo:str,compra_numero:int,request:dict):
    local=clean_text(local_codigo,10) or ""; doc_type=clean_text(compra_tipo,2) or ""; provider=clean_text(proveedor_codigo,20) or ""
    with cursor() as (_,cur):
        result = close_purchase_inventory(cur, local, doc_type, provider, compra_numero, request.get("user_id", 0))
    return json_safe(result)


REPORT_DEFINITIONS = {
    "estadisticas": {"title": "Estadisticas de ventas", "description": "Evolucion diaria de documentos, neto, impuestos, ventas y pagos.", "columns": [("periodo","Dia"),("documentos","Documentos"),("neto","Neto"),("iva","IVA"),("ila","ILA"),("total","Total"),("pagado","Pagado"),("saldo","Saldo")], "money": ["neto","iva","ila","total","pagado","saldo"], "chart": "total"},
    "vendedores": {"title": "Ventas por vendedor", "description": "Rendimiento comercial, cobranza y ticket promedio por vendedor.", "columns": [("codigo","Codigo"),("nombre","Vendedor"),("documentos","Documentos"),("clientes","Clientes"),("neto","Neto"),("total","Total"),("pagado","Pagado"),("saldo","Saldo"),("ticket","Ticket promedio")], "money": ["neto","total","pagado","saldo","ticket"], "chart": "total", "chart_label": "nombre"},
    "rutas": {"title": "Ventas por rutas", "description": "Actividad comercial y saldos pendientes por ruta de despacho.", "columns": [("codigo","Ruta"),("nombre","Descripcion"),("documentos","Documentos"),("clientes","Clientes"),("total","Total"),("pagado","Pagado"),("saldo","Saldo"),("ticket","Ticket promedio")], "money": ["total","pagado","saldo","ticket"], "chart": "total", "chart_label": "nombre"},
    "familias": {"title": "Ventas por familias", "description": "Unidades, neto y participación comercial por familia de productos.", "columns": [("codigo","Codigo"),("nombre","Familia"),("productos","Productos"),("cantidad","Cantidad"),("neto","Neto"),("impuestos","Impuestos"),("total","Total")], "money": ["neto","impuestos","total"], "chart": "total", "chart_label": "nombre"},
    "formas-pago": {"title": "Ventas por formas de pago", "description": "Recaudacion agrupada por medios de pago utilizados.", "columns": [("codigo","Codigo"),("nombre","Forma de pago"),("operaciones","Operaciones"),("documentos","Documentos"),("monto","Monto"),("participacion","Participacion %")], "money": ["monto"], "chart": "monto", "chart_label": "nombre"},
    "pendientes": {"title": "Facturas pendientes", "description": "Facturas con saldo, antiguedad y responsables de cobranza.", "columns": [("fecha","Fecha"),("folio","Folio"),("rut","RUT"),("cliente","Cliente"),("vendedor","Vendedor"),("ruta","Ruta"),("total","Total"),("pagado","Pagado"),("saldo","Saldo"),("dias","Dias vencidos")], "money": ["total","pagado","saldo"], "chart": "saldo"},
    "cta-cte": {"title": "Cuentas corrientes de clientes", "description": "Deuda, documentos abiertos y antiguedad por cliente.", "columns": [("rut","RUT"),("cliente","Cliente"),("ruta","Ruta"),("vendedor","Vendedor"),("documentos","Documentos"),("deuda","Deuda sistema"),("saldo_documentos","Saldo documentos"),("ultimo_documento","Ultima venta")], "money": ["deuda","saldo_documentos"], "chart": "saldo_documentos"},
    "cobros": {"title": "Cobros y recaudacion", "description": "Comprobantes de pago y recaudacion por fecha y forma de pago.", "columns": [("fecha","Fecha"),("comprobantes","Comprobantes"),("clientes","Clientes"),("operaciones","Operaciones"),("monto","Monto cobrado")], "money": ["monto"], "chart": "monto"},
    "cartola": {"title": "Cartola de clientes", "description": "Documentos, pagos y saldo del cliente seleccionado.", "columns": [("fecha","Fecha"),("tipo","Tipo"),("numero","Documento"),("detalle","Detalle"),("cargo","Cargo"),("abono","Abono"),("saldo","Saldo")], "money": ["cargo","abono","saldo"], "chart": "cargo", "needs_client": True},
    "compras": {"title": "Compras por proveedor", "description": "Documentos, neto, impuestos, pagos y saldo por proveedor.", "columns": [("codigo","Codigo"),("proveedor","Proveedor"),("documentos","Documentos"),("neto","Neto"),("impuestos","Impuestos"),("total","Total"),("pagado","Pagado"),("saldo","Saldo")], "money": ["neto","impuestos","total","pagado","saldo"], "chart": "total"},
    "inventario": {"title": "Inventario valorizado", "description": "Stock actual estimado, costo y valorizacion por producto.", "columns": [("codigo","Codigo"),("producto","Producto"),("familia","Familia"),("unidad","Unidad"),("stock","Stock"),("costo","Costo"),("venta","Precio venta"),("valorizado","Valorizado costo")], "money": ["costo","venta","valorizado"], "chart": "valorizado"},
    "stock-bajo": {"title": "Stock critico y reposicion", "description": "Productos bajo minimo y sugerencia de reposicion.", "columns": [("codigo","Codigo"),("producto","Producto"),("familia","Familia"),("stock","Stock"),("minimo","Stock minimo"),("faltante","Faltante"),("costo_reposicion","Costo reposicion")], "money": ["costo_reposicion"], "chart": "faltante"},
    "productos": {"title": "Ranking de productos", "description": "Productos con mayor rotacion, unidades y venta acumulada.", "columns": [("codigo","Codigo"),("producto","Producto"),("familia","Familia"),("documentos","Documentos"),("cantidad","Cantidad"),("neto","Neto"),("total","Total")], "money": ["neto","total"], "chart": "total"},
}


@app.get("/api/reportes/catalogos")
def reportes_catalogos():
    with cursor() as (_, cur):
        cur.execute("SELECT vendedor_codigo AS value,vendedor_nombre AS label FROM vendedores WHERE COALESCE(vendedor_estado,'A')<>'B' ORDER BY vendedor_nombre")
        sellers = clean_rows(cur.fetchall())
        cur.execute("SELECT CAST(ruta_id AS CHAR) AS value,ruta_nombre AS label FROM rutas ORDER BY ruta_nombre")
        routes = clean_rows(cur.fetchall())
        cur.execute("SELECT familia_codigo AS value,familia_descripcion AS label FROM familias ORDER BY familia_descripcion")
        families = clean_rows(cur.fetchall())
    return json_safe({"vendedores": sellers, "rutas": routes, "familias": families, "reportes": [{"id": key, **value} for key,value in REPORT_DEFINITIONS.items()]})


def report_filters(alias, fecha_desde, fecha_hasta, local_codigo):
    filters = []; params = []
    if fecha_desde: filters.append(f"{alias}.venta_fecha>=%s"); params.append(parse_date(fecha_desde))
    if fecha_hasta: filters.append(f"{alias}.venta_fecha<=%s"); params.append(parse_date(fecha_hasta))
    local = clean_text(local_codigo, 10)
    if local: filters.append(f"{alias}.local_codigo=%s"); params.append(local)
    return filters, params


@app.get("/api/reportes/{report_id}")
def ejecutar_reporte(
    report_id: str, fecha_desde: str | None = Query(None), fecha_hasta: str | None = Query(None),
    local_codigo: str | None = Query(None), vendedor_codigo: str | None = Query(None),
    ruta_id: str | None = Query(None), familia_codigo: str | None = Query(None),
    cliente_rut: str | None = Query(None), limit: int = Query(300, ge=1, le=1000),
):
    definition = REPORT_DEFINITIONS.get(report_id)
    if not definition: raise HTTPException(404, "Reporte no disponible")
    filters, params = report_filters("v", fecha_desde, fecha_hasta, local_codigo)
    filters.append("v.venta_tipo IN ('FE','FA','BO','NC')")
    seller = clean_text(vendedor_codigo, 10); route = clean_text(ruta_id, 20); family = clean_text(familia_codigo, 20); client = clean_text(cliente_rut, 20)
    if seller: filters.append("v.vendedor_codigo=%s"); params.append(seller)
    if route: filters.append("CAST(c.ruta_id AS CHAR)=%s"); params.append(route)
    where = " AND ".join(filters)
    rows = []
    with cursor() as (_, cur):
        if report_id == "estadisticas":
            cur.execute(f"""SELECT DATE_FORMAT(v.venta_fecha,'%Y-%m-%d') periodo,COUNT(*) documentos,SUM(COALESCE(v.venta_neto1,0)) neto,SUM(COALESCE(v.venta_iva1,0)) iva,SUM(COALESCE(v.venta_ila1,0)) ila,SUM(COALESCE(v.venta_totalventa,0)) total,SUM(COALESCE(v.venta_pagototal,0)) pagado,SUM(COALESCE(v.venta_totalventa,0)-COALESCE(v.venta_pagototal,0)) saldo FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut WHERE {where} GROUP BY v.venta_fecha ORDER BY v.venta_fecha""", tuple(params)); rows=clean_rows(cur.fetchall())
        elif report_id == "vendedores":
            cur.execute(f"""SELECT v.vendedor_codigo codigo,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) nombre,COUNT(*) documentos,COUNT(DISTINCT v.cliente_rut) clientes,SUM(COALESCE(v.venta_neto1,0)) neto,SUM(COALESCE(v.venta_totalventa,0)) total,SUM(COALESCE(v.venta_pagototal,0)) pagado,SUM(COALESCE(v.venta_totalventa,0)-COALESCE(v.venta_pagototal,0)) saldo,ROUND(AVG(COALESCE(v.venta_totalventa,0)),0) ticket FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo WHERE {where} GROUP BY v.vendedor_codigo,ve.vendedor_nombre ORDER BY total DESC LIMIT %s""", tuple(params+[limit])); rows=clean_rows(cur.fetchall())
        elif report_id == "rutas":
            cur.execute(f"""SELECT CAST(c.ruta_id AS CHAR) codigo,COALESCE(r.ruta_nombre,'SIN RUTA') nombre,COUNT(*) documentos,COUNT(DISTINCT v.cliente_rut) clientes,SUM(COALESCE(v.venta_totalventa,0)) total,SUM(COALESCE(v.venta_pagototal,0)) pagado,SUM(COALESCE(v.venta_totalventa,0)-COALESCE(v.venta_pagototal,0)) saldo,ROUND(AVG(COALESCE(v.venta_totalventa,0)),0) ticket FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN rutas r ON r.ruta_id=c.ruta_id WHERE {where} GROUP BY c.ruta_id,r.ruta_nombre ORDER BY total DESC LIMIT %s""", tuple(params+[limit])); rows=clean_rows(cur.fetchall())
        elif report_id in ("familias","productos"):
            detail_filters=list(filters); detail_params=list(params)
            if family: detail_filters.append("p.familia_codigo=%s"); detail_params.append(family)
            detail_where=" AND ".join(detail_filters)
            if report_id=="familias": sql=f"""SELECT p.familia_codigo codigo,COALESCE(f.familia_descripcion,'SIN FAMILIA') nombre,COUNT(DISTINCT d.producto_codigo) productos,SUM(COALESCE(d.venta_cantidad,0)) cantidad,SUM(COALESCE(d.venta_lineaneto,0)) neto,SUM(COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0)) impuestos,SUM(COALESCE(d.venta_lineaneto,0)+COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0)) total FROM ventaslevel2 d INNER JOIN ventas v ON v.venta_numero=d.venta_numero AND v.venta_tipo=d.venta_tipo AND v.local_codigo=d.local_codigo LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo LEFT JOIN familias f ON f.familia_codigo=p.familia_codigo WHERE {detail_where} GROUP BY p.familia_codigo,f.familia_descripcion ORDER BY total DESC LIMIT %s"""
            else: sql=f"""SELECT d.producto_codigo codigo,COALESCE(p.producto_descripcion,d.venta_descripcion) producto,COALESCE(f.familia_descripcion,'SIN FAMILIA') familia,COUNT(DISTINCT CONCAT(d.local_codigo,'-',d.venta_tipo,'-',d.venta_numero)) documentos,SUM(COALESCE(d.venta_cantidad,0)) cantidad,SUM(COALESCE(d.venta_lineaneto,0)) neto,SUM(COALESCE(d.venta_lineaneto,0)+COALESCE(d.venta_lineaiva,0)+COALESCE(d.venta_lineaila,0)) total FROM ventaslevel2 d INNER JOIN ventas v ON v.venta_numero=d.venta_numero AND v.venta_tipo=d.venta_tipo AND v.local_codigo=d.local_codigo LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN productos p ON p.producto_codigo=d.producto_codigo LEFT JOIN familias f ON f.familia_codigo=p.familia_codigo WHERE {detail_where} GROUP BY d.producto_codigo,p.producto_descripcion,d.venta_descripcion,f.familia_descripcion ORDER BY total DESC LIMIT %s"""
            cur.execute(sql,tuple(detail_params+[limit])); rows=clean_rows(cur.fetchall())
        elif report_id == "formas-pago":
            cur.execute(f"""SELECT p.fpago_codigo codigo,COALESCE(fp.fpago_descripcion,p.fpago_codigo) nombre,COUNT(*) operaciones,COUNT(DISTINCT CONCAT(p.local_codigo,'-',p.venta_tipo,'-',p.venta_numero)) documentos,SUM(COALESCE(p.venta_pagomonto,0)) monto,0 participacion FROM ventaslevel1 p INNER JOIN ventas v ON v.venta_numero=p.venta_numero AND v.venta_tipo=p.venta_tipo AND v.local_codigo=p.local_codigo LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN formasdepago fp ON fp.fpago_codigo=p.fpago_codigo WHERE {where} GROUP BY p.fpago_codigo,fp.fpago_descripcion ORDER BY monto DESC""",tuple(params)); rows=clean_rows(cur.fetchall()); total=sum(to_float(x.get('monto'),0) for x in rows); [x.update(participacion=round(to_float(x.get('monto'),0)*100/total,2) if total else 0) for x in rows]
        elif report_id == "pendientes":
            pending_filters=[x for x in filters if "venta_tipo" not in x]; pending_filters.append("v.venta_tipo IN ('FE','FA')"); pending_filters.append("COALESCE(v.venta_totalventa,0)>COALESCE(v.venta_pagototal,0)")
            cur.execute(f"""SELECT v.venta_fecha fecha,COALESCE(v.venta_folio,v.venta_numero) folio,v.cliente_rut rut,COALESCE(c.cliente_nombre,'SIN CLIENTE') cliente,COALESCE(ve.vendedor_nombre,v.vendedor_codigo) vendedor,COALESCE(r.ruta_nombre,'SIN RUTA') ruta,v.venta_totalventa total,COALESCE(v.venta_pagototal,0) pagado,v.venta_totalventa-COALESCE(v.venta_pagototal,0) saldo,GREATEST(DATEDIFF(CURDATE(),COALESCE(v.venta_fechavto,v.venta_fecha)),0) dias FROM ventas v LEFT JOIN clientes c ON c.cliente_rut=v.cliente_rut LEFT JOIN vendedores ve ON ve.vendedor_codigo=v.vendedor_codigo LEFT JOIN rutas r ON r.ruta_id=c.ruta_id WHERE {' AND '.join(pending_filters)} ORDER BY dias DESC,saldo DESC LIMIT %s""",tuple(params+[limit])); rows=clean_rows(cur.fetchall())
        elif report_id == "cta-cte":
            current_filters=[x for x in filters if "venta_tipo" not in x]; current_filters.append("v.venta_tipo IN ('FE','FA')")
            cur.execute(f"""SELECT c.cliente_rut rut,c.cliente_nombre cliente,COALESCE(r.ruta_nombre,'SIN RUTA') ruta,COALESCE(ve.vendedor_nombre,c.cliente_vendedor) vendedor,COUNT(CASE WHEN COALESCE(v.venta_totalventa,0)>COALESCE(v.venta_pagototal,0) THEN 1 END) documentos,COALESCE(c.cliente_deudatotal,0) deuda,SUM(GREATEST(COALESCE(v.venta_totalventa,0)-COALESCE(v.venta_pagototal,0),0)) saldo_documentos,MAX(v.venta_fecha) ultimo_documento FROM clientes c LEFT JOIN ventas v ON v.cliente_rut=c.cliente_rut LEFT JOIN rutas r ON r.ruta_id=c.ruta_id LEFT JOIN vendedores ve ON ve.vendedor_codigo=c.cliente_vendedor WHERE {' AND '.join(current_filters)} GROUP BY c.cliente_rut,c.cliente_nombre,r.ruta_nombre,ve.vendedor_nombre,c.cliente_vendedor,c.cliente_deudatotal HAVING saldo_documentos>0 OR deuda>0 ORDER BY saldo_documentos DESC LIMIT %s""",tuple(params+[limit])); rows=clean_rows(cur.fetchall())
        elif report_id == "cobros":
            cob_filters=[]; cob_params=[]
            if fecha_desde: cob_filters.append("cp.comprobante_fecha>=%s"); cob_params.append(parse_date(fecha_desde))
            if fecha_hasta: cob_filters.append("cp.comprobante_fecha<=%s"); cob_params.append(parse_date(fecha_hasta))
            if local_codigo: cob_filters.append("cp.local_codigo=%s"); cob_params.append(clean_text(local_codigo,10))
            sql_where=" AND ".join(cob_filters) or "1=1"
            cur.execute(f"""SELECT cp.comprobante_fecha fecha,COUNT(DISTINCT CONCAT(cp.local_codigo,'-',cp.comprobante_numero)) comprobantes,COUNT(DISTINCT cp.cliente_rut) clientes,COUNT(*) operaciones,SUM(COALESCE(p.comprobante_monto,0)) monto FROM comprobantepago cp INNER JOIN comprobantepagolevel1 p ON p.comprobante_numero=cp.comprobante_numero AND p.local_codigo=cp.local_codigo WHERE {sql_where} GROUP BY cp.comprobante_fecha ORDER BY cp.comprobante_fecha""",tuple(cob_params)); rows=clean_rows(cur.fetchall())
        elif report_id == "cartola":
            if not client: rows=[]
            else:
                cart_filters=["v.cliente_rut=%s","v.venta_tipo IN ('FE','FA','NC')"]; cart_params=[client]
                if fecha_desde: cart_filters.append("v.venta_fecha>=%s"); cart_params.append(parse_date(fecha_desde))
                if fecha_hasta: cart_filters.append("v.venta_fecha<=%s"); cart_params.append(parse_date(fecha_hasta))
                if local_codigo: cart_filters.append("v.local_codigo=%s"); cart_params.append(clean_text(local_codigo,10))
                cur.execute(f"""SELECT v.venta_fecha fecha,v.venta_tipo tipo,COALESCE(v.venta_folio,v.venta_numero) numero,CONCAT('Documento ',v.venta_tipo) detalle,CASE WHEN v.venta_tipo='NC' THEN 0 ELSE v.venta_totalventa END cargo,CASE WHEN v.venta_tipo='NC' THEN v.venta_totalventa ELSE COALESCE(v.venta_pagototal,0) END abono,0 saldo FROM ventas v WHERE {' AND '.join(cart_filters)} ORDER BY v.venta_fecha,v.venta_numero""",tuple(cart_params)); rows=clean_rows(cur.fetchall()); balance=0
                for row in rows: balance+=to_float(row.get('cargo'),0)-to_float(row.get('abono'),0); row['saldo']=balance
        elif report_id == "compras":
            buy_filters=[]; buy_params=[]
            if fecha_desde: buy_filters.append("c.compra_fecha>=%s"); buy_params.append(parse_date(fecha_desde))
            if fecha_hasta: buy_filters.append("c.compra_fecha<=%s"); buy_params.append(parse_date(fecha_hasta))
            if local_codigo: buy_filters.append("c.local_codigo=%s"); buy_params.append(clean_text(local_codigo,10))
            buy_where=" AND ".join(buy_filters) or "1=1"
            cur.execute(f"""SELECT c.proveedor_codigo codigo,COALESCE(p.proveedor_nombre,c.proveedor_codigo) proveedor,COUNT(*) documentos,SUM(COALESCE(c.compra_neto,0)) neto,SUM(COALESCE(c.compra_iva,0)+COALESCE(c.compra_ila,0)) impuestos,SUM(COALESCE(c.compra_totalcompra,0)) total,COALESCE((SELECT SUM(pc.compra_pagomonto) FROM pagocompraslevel2 pc WHERE pc.proveedor_codigo=c.proveedor_codigo AND pc.compra_pagofecha BETWEEN %s AND %s),0) pagado,0 saldo FROM compras c LEFT JOIN proveedores p ON p.proveedor_codigo=c.proveedor_codigo WHERE {buy_where} GROUP BY c.proveedor_codigo,p.proveedor_nombre ORDER BY total DESC LIMIT %s""",tuple([parse_date(fecha_desde or '2000-01-01'),parse_date(fecha_hasta or str(datetime.date.today()))]+buy_params+[limit])); rows=clean_rows(cur.fetchall()); [x.update(saldo=to_float(x.get('total'),0)-to_float(x.get('pagado'),0)) for x in rows]
        elif report_id in ("inventario","stock-bajo"):
            stock_expr="COALESCE((SELECT SUM(CASE WHEN sm.stock_tm='I' THEN sm.stock_cantidad WHEN sm.stock_tm='E' THEN -sm.stock_cantidad ELSE sm.stock_cantidad END) FROM stock_mensualeslevel1 sm WHERE sm.producto_codigo=p.producto_codigo),0)"
            product_filter="WHERE COALESCE(p.producto_estado,'A') NOT IN ('I','N')"; product_params=[]
            if family: product_filter+=" AND p.familia_codigo=%s"; product_params.append(family)
            if report_id=="inventario": sql=f"""SELECT p.producto_codigo codigo,p.producto_descripcion producto,COALESCE(f.familia_descripcion,'SIN FAMILIA') familia,p.unidad_codigo unidad,{stock_expr} stock,COALESCE(p.producto_costosinflete,p.producto_costo,0) costo,COALESCE(p.producto_venta,0) venta,ROUND({stock_expr}*COALESCE(p.producto_costosinflete,p.producto_costo,0),0) valorizado FROM productos p LEFT JOIN familias f ON f.familia_codigo=p.familia_codigo {product_filter} ORDER BY valorizado DESC LIMIT %s"""
            else: sql=f"""SELECT p.producto_codigo codigo,p.producto_descripcion producto,COALESCE(f.familia_descripcion,'SIN FAMILIA') familia,{stock_expr} stock,COALESCE(p.producto_stockmin,0) minimo,GREATEST(COALESCE(p.producto_stockmin,0)-({stock_expr}),0) faltante,ROUND(GREATEST(COALESCE(p.producto_stockmin,0)-({stock_expr}),0)*COALESCE(p.producto_costosinflete,p.producto_costo,0),0) costo_reposicion FROM productos p LEFT JOIN familias f ON f.familia_codigo=p.familia_codigo {product_filter} HAVING faltante>0 ORDER BY faltante DESC LIMIT %s"""
            cur.execute(sql,tuple(product_params+[limit])); rows=clean_rows(cur.fetchall())
    totals={}
    for key in definition.get("money",[]): totals[key]=sum(to_float(row.get(key),0) for row in rows)
    return json_safe({"definition": definition, "rows": rows, "totals": totals, "count": len(rows)})


@app.get("/api/pos/producto")
def pos_producto(codigo: str = Query(...), local_codigo: str = Query(...)):
    local = clean_text(local_codigo, 10) or ""
    with cursor() as (_, cur):
        product = pos_product(cur, codigo, local)
    if not product:
        raise HTTPException(404, "Producto o codigo de barra no encontrado")
    return json_safe(product)


@app.get("/api/pos/productos")
def pos_productos(q: str = Query(...), local_codigo: str = Query(...), limit: int = Query(20, ge=1, le=50)):
    term = clean_text(q, 80) or ""
    price_list = pos_price_list(clean_text(local_codigo, 10) or "")
    with cursor() as (_, cur):
        cur.execute("""
            SELECT p.producto_codigo, p.producto_descripcion, p.producto_barra, p.unidad_codigo,
                   COALESCE(pl.lista_venta, p.producto_venta, 0) AS precio_venta
            FROM productos p
            LEFT JOIN precioslevel1 pl ON pl.producto_codigo = p.producto_codigo AND pl.lista_codigo = %s
            WHERE COALESCE(p.producto_estado, 'A') NOT IN ('I','N')
              AND (p.producto_codigo LIKE %s OR p.producto_descripcion LIKE %s OR p.producto_barra LIKE %s)
            ORDER BY p.producto_descripcion LIMIT %s
        """, (price_list, f"{term}%", f"%{term}%", f"%{term}%", limit))
        rows = clean_rows(cur.fetchall())
    return json_safe(rows)


@app.post("/api/pos/boletas")
def pos_emitir_boleta(payload: PosSaleRequest):
    local_codigo = clean_text(payload.local_codigo, 10) or ""
    caja_codigo = to_int(payload.caja_codigo, 0)
    user_id = to_int(payload.user_id, 0)
    if not local_codigo or caja_codigo <= 0 or user_id <= 0:
        raise HTTPException(400, "Faltan local, caja o usuario")
    if not payload.lines:
        raise HTTPException(400, "Agregue al menos un producto")
    if not payload.payments:
        raise HTTPException(400, "Ingrese una forma de pago")

    today = datetime.date.today()
    now = datetime.datetime.now()
    with cursor() as (cn, cur):
        ensure_app_security(cur)
        cur.execute("SELECT 1 FROM app_caja_cierres WHERE local_codigo=%s AND caja_codigo=%s AND cierre_fecha=%s AND estado='C'", (local_codigo,caja_codigo,today))
        if cur.fetchone():
            raise HTTPException(409, "La caja esta cerrada para la fecha actual")
        cur.execute("SELECT local_boleta, local_bodega FROM locales WHERE local_codigo = %s FOR UPDATE", (local_codigo,))
        local = clean_row(cur.fetchone() or {})
        if not local:
            raise HTTPException(404, "Local no encontrado")
        cur.execute("SELECT 1 FROM caja WHERE local_codigo = %s AND caja_codigo = %s LIMIT 1", (local_codigo, caja_codigo))
        if not cur.fetchone():
            raise HTTPException(400, "Caja no valida para el local")
        folio = to_int(local.get("local_boleta"), 0)
        if folio <= 0:
            raise HTTPException(400, "El local no tiene folio de boleta configurado")
        cur.execute("SELECT 1 FROM ventas WHERE venta_numero = %s AND venta_tipo = 'BO' AND local_codigo = %s", (folio, local_codigo))
        if cur.fetchone():
            raise HTTPException(409, "El folio de boleta ya existe; revise el contador del local")

        consolidated = {}
        for requested in payload.lines:
            product = pos_product(cur, requested.producto_codigo, local_codigo)
            if not product:
                raise HTTPException(400, f"Producto {requested.producto_codigo} no encontrado")
            code = clean_text(product.get("producto_codigo"), 20) or ""
            quantity = decimal_value(requested.cantidad)
            discount = max(decimal.Decimal("0"), min(decimal.Decimal("100"), decimal_value(requested.descuento)))
            if quantity <= 0:
                raise HTTPException(400, f"Cantidad invalida para {code}")
            if code in consolidated:
                consolidated[code]["cantidad"] += quantity
                consolidated[code]["descuento"] = discount
            else:
                consolidated[code] = {"product": product, "cantidad": quantity, "descuento": discount}

        total_neto = total_iva = total_ila = total_venta = 0
        details = []
        for code, item in consolidated.items():
            product = item["product"]
            quantity = item["cantidad"]
            factor = decimal.Decimal("1") - item["descuento"] / decimal.Decimal("100")
            line_neto = round_business(decimal_value(product.get("precio_neto")) * quantity * factor)
            line_iva = round_business(decimal_value(product.get("precio_iva")) * quantity * factor)
            line_ila = round_business(decimal_value(product.get("precio_ila")) * quantity * factor)
            line_total = round_business(decimal_value(product.get("precio_venta")) * quantity * factor)
            total_neto += line_neto
            total_iva += line_iva
            total_ila += line_ila
            total_venta += line_total
            details.append((code, product, quantity, item["descuento"], line_neto, line_iva, line_ila, line_total))

        payments = []
        tendered = 0
        for payment in payload.payments:
            code = clean_text(payment.fpago_codigo, 10) or ""
            amount = max(0, to_int(payment.monto, 0))
            if not code or amount <= 0:
                continue
            cur.execute("SELECT 1 FROM formasdepago WHERE fpago_codigo = %s AND COALESCE(fpago_activo, 'S') <> 'N'", (code,))
            if not cur.fetchone():
                raise HTTPException(400, f"Forma de pago {code} no valida")
            tendered += amount
            payments.append({"code": code, "amount": amount, "document": clean_text(payment.numero_documento, 20) or ""})
        cash = next((payment for payment in payments if payment["code"] == "01"), None)
        non_cash = sum(payment["amount"] for payment in payments if payment["code"] != "01")
        cash_due_exact = max(0, total_venta - non_cash)
        cash_due = chile_cash_round(cash_due_exact) if cash else cash_due_exact
        payable_total = non_cash + cash_due
        if tendered < payable_total:
            raise HTTPException(400, f"Faltan {payable_total - tendered} por pagar")
        change = tendered - payable_total
        if change > 0 and (not cash or cash["amount"] < change):
            raise HTTPException(400, "El vuelto solo puede descontarse del pago en efectivo")
        if cash:
            cash["applied"] = cash["amount"] - change
        for payment in payments:
            payment.setdefault("applied", payment["amount"])

        seller = clean_text(payload.vendedor_codigo, 10) or ""
        client = clean_text(payload.cliente_rut, 10) or "0"
        payment_label = payments[0]["code"] if len(payments) == 1 else "MULTI"
        card_total = sum(p["applied"] for p in payments if p["code"] in ("05", "06"))
        cur.execute("""
            INSERT INTO ventas (
                venta_numero, venta_tipo, local_codigo, venta_fecha, cliente_rut,
                venta_ult, venta_mes, venta_ano, venta_estado, venta_tarjeta,
                vendedor_codigo, venta_descuento, venta_totalventa, caja_codigo,
                venta_turno, venta_hora, venta_condicion, venta_pago, venta_impreso,
                venta_impresion, venta_folio, venta_usuario, venta_neto1, venta_iva1,
                venta_ila1, venta_totalchequesprot, venta_pagototal
            ) VALUES (
                %s, 'BO', %s, %s, %s, DAY(%s), MONTH(%s), YEAR(%s), 'A', %s,
                %s, 0, %s, %s, 1, %s, %s, %s, 'N', 'N', %s, %s, %s, %s, %s, 0, %s
            )
        """, (
            folio, local_codigo, today, client, today, today, today, card_total,
            seller, total_venta, caja_codigo, now, payment_label, payment_label,
            folio, user_id, total_neto, total_iva, total_ila, total_venta,
        ))
        bodega = clean_text(local.get("local_bodega"), 10) or local_codigo
        for code, product, quantity, discount, neto, iva, ila, total in details:
            unit_net = to_int(product.get("precio_neto"), 0)
            cur.execute("""
                INSERT INTO ventaslevel2 (
                    venta_numero, venta_tipo, local_codigo, producto_codigo, bodega_codigo,
                    venta_lineaila, venta_lineaiva, venta_lineaneto, venta_precio,
                    venta_cantidad, empleado_codigo, venta_descuentol, venta_unidadenvase,
                    venta_descripcion, venta_precioventa, venta_totalneto, venta_cajas
                ) VALUES (%s, 'BO', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)
            """, (
                folio, local_codigo, code, bodega, ila, iva, neto, unit_net,
                float(quantity), seller, float(discount), to_float(product.get("producto_unidadenvase"), 0),
                clean_text(product.get("producto_descripcion"), 50) or code, unit_net, neto,
            ))
        for index, payment in enumerate(payments, start=1):
            cur.execute("""
                INSERT INTO ventaslevel1 (
                    venta_numero, venta_tipo, local_codigo, venta_pagoitem, fpago_codigo,
                    venta_pagofecha, banco_codigo, venta_pagomonto, venta_numerodoc,
                    venta_pagado, venta_pagomontop, venta_pagooc, venta_chequeprotestado, venta_pagoreal
                ) VALUES (%s, 'BO', %s, %s, %s, %s, '', %s, %s, 'S', %s, '', 'N', %s)
            """, (
                folio, local_codigo, index, payment["code"], today, payment["applied"],
                payment["document"], payment["applied"], payment["amount"],
            ))
        cur.execute("UPDATE locales SET local_boleta = %s WHERE local_codigo = %s", (folio + 1, local_codigo))
        if payload.validate_only:
            cn.rollback()
            return json_safe({"ok": True, "validated": True, "folio": folio, "total": total_venta, "total_pagar": payable_total, "redondeo": payable_total-total_venta, "vuelto": change})
    return json_safe({"ok": True, "folio": folio, "total": total_venta, "total_pagar": payable_total, "redondeo": payable_total-total_venta, "vuelto": change})


@app.get("/api/rutas")
def rutas():
    with cursor() as (_, cur):
        try:
            cur.execute("""
                SELECT CAST(ruta_id AS CHAR) AS ruta_id, ruta_nombre
                FROM rutas
                ORDER BY ruta_nombre
            """)
            return json_safe(clean_rows(cur.fetchall()))
        except Exception:
            return json_safe([])


@app.get("/api/vendedores")
def vendedores():
    with cursor() as (_, cur):
        cols = table_columns(cur, "vendedores")
        name_col = next((c for c in ("vendedor_nombre", "vendedor_descripcion", "vendedor_name", "nombre") if c in cols), None)
        name_expr = f"COALESCE({name_col}, CAST(vendedor_codigo AS CHAR))" if name_col else "CAST(vendedor_codigo AS CHAR)"
        cur.execute(f"""
            SELECT CAST(vendedor_codigo AS CHAR) AS vendedor_codigo, {name_expr} AS vendedor_nombre
            FROM vendedores
            ORDER BY {name_expr}
        """)
        return json_safe(clean_rows(cur.fetchall()))


@app.get("/api/maestros/{maestro}")
def maestro_listado(
    maestro: str,
    q: str | None = Query(None),
    limit: int = Query(150, ge=1, le=500),
):
    config = MAESTROS.get(clean_text(maestro, 40) or "")
    if not config:
        raise HTTPException(404, "Mantenedor no disponible")

    fields = config["fields"]
    q = clean_text(q, 100)
    select_sql = ", ".join(f"`{field}`" for field, _ in fields)
    where_sql = ""
    params = []
    if q:
        searchable = " OR ".join(
            f"COALESCE(CAST(`{field}` AS CHAR), '') LIKE %s" for field, _ in fields
        )
        where_sql = f"WHERE ({searchable})"
        params.extend([f"%{q}%"] * len(fields))

    with cursor() as (_, cur):
        cur.execute(
            f"SELECT {select_sql} FROM `{config['table']}` "
            f"{where_sql} ORDER BY {config['order']} LIMIT %s",
            tuple(params + [limit]),
        )
        rows = clean_rows(cur.fetchall())

    return json_safe({
        "meta": {
            "id": maestro,
            "title": config["title"],
            "table": config["table"],
            "key": config["key"],
            "name": config["name"],
            "fields": [{"key": field, "label": label} for field, label in fields],
        },
        "rows": rows,
    })


@app.post("/api/maestros/{maestro}")
def maestro_crear(maestro: str, payload: SaveRequest):
    config = MAESTROS.get(clean_text(maestro, 40) or "")
    if not config:
        raise HTTPException(404, "Mantenedor no disponible")
    allowed = {field for field, _ in config["fields"]}
    data = {key: value for key, value in payload.data.items() if key in allowed}
    if maestro == "usuarios" and data.get("SecUserId") in (None, ""):
        data.pop("SecUserId", None)
    if not data:
        raise HTTPException(400, "No hay datos para guardar")
    with cursor() as (_, cur):
        columns = table_columns(cur, config["table"])
        data = {key: value for key, value in data.items() if key in columns}
        keys = ", ".join(f"`{key}`" for key in data)
        marks = ", ".join(["%s"] * len(data))
        cur.execute(f"INSERT INTO `{config['table']}` ({keys}) VALUES ({marks})", tuple(data.values()))
    return json_safe({"ok": True})


@app.put("/api/maestros/{maestro}")
def maestro_actualizar(maestro: str, payload: SaveRequest):
    config = MAESTROS.get(clean_text(maestro, 40) or "")
    if not config:
        raise HTTPException(404, "Mantenedor no disponible")
    allowed = {field for field, _ in config["fields"]}
    data = {key: value for key, value in payload.data.items() if key in allowed}
    original = payload.original or {}
    key_fields = [config["key"]]
    if maestro == "cajas":
        key_fields = ["caja_fecha", "local_codigo", "caja_codigo"]
    if any(original.get(key) in (None, "") for key in key_fields):
        raise HTTPException(400, "Falta la clave original del registro")
    with cursor() as (_, cur):
        columns = table_columns(cur, config["table"])
        data = {key: value for key, value in data.items() if key in columns}
        set_sql = ", ".join(f"`{key}` = %s" for key in data)
        where_sql = " AND ".join(f"`{key}` = %s" for key in key_fields)
        cur.execute(
            f"UPDATE `{config['table']}` SET {set_sql} WHERE {where_sql}",
            tuple(data.values()) + tuple(original[key] for key in key_fields),
        )
        if cur.rowcount == 0:
            cur.execute(f"SELECT 1 FROM `{config['table']}` WHERE {where_sql} LIMIT 1", tuple(original[key] for key in key_fields))
            if not cur.fetchone():
                raise HTTPException(404, "Registro no encontrado")
    return json_safe({"ok": True})


@app.get("/api/catalogos/clientes")
def catalogos_clientes():
    with cursor() as (_, cur):
        out = {
            "rutas": [],
            "ciudades": [],
            "comunas": [],
            "vendedores": [],
            "listas": [],
            "condiciones": [],
            "estados": [{"value": "A", "label": "ACTIVO"}, {"value": "I", "label": "INACTIVO"}],
        }
        try:
            cur.execute("SELECT CAST(ruta_id AS CHAR) AS value, ruta_nombre AS label FROM rutas ORDER BY ruta_nombre")
            out["rutas"] = clean_rows(cur.fetchall())
        except Exception:
            pass
        try:
            cur.execute("SELECT ciudad_codigo AS value, ciudad_codigo AS label FROM ciudades ORDER BY ciudad_codigo")
            out["ciudades"] = clean_rows(cur.fetchall())
        except Exception:
            pass
        try:
            cur.execute("SELECT Ciudad_codigo AS ciudad, Comuna AS value, Comuna AS label FROM ciudadeslevel1 ORDER BY Ciudad_codigo, Comuna")
            out["comunas"] = clean_rows(cur.fetchall())
        except Exception:
            pass
        try:
            cur.execute("SELECT vendedor_codigo AS value, vendedor_nombre AS label FROM vendedores ORDER BY vendedor_nombre")
            out["vendedores"] = clean_rows(cur.fetchall())
        except Exception:
            pass
        try:
            cur.execute("SELECT lista_codigo AS value, COALESCE(lista_descripcion, lista_codigo) AS label FROM listaprecios ORDER BY lista_codigo")
            out["listas"] = clean_rows(cur.fetchall())
        except Exception:
            pass
        try:
            cur.execute("SELECT fpago_codigo AS value, COALESCE(fpago_descripcion, fpago_codigo) AS label FROM formasdepago ORDER BY fpago_descripcion")
            out["condiciones"] = clean_rows(cur.fetchall())
        except Exception:
            out["condiciones"] = [
                {"value": "EFECTIVO", "label": "EFECTIVO"},
                {"value": "CREDITO 30 DIAS", "label": "CREDITO 30 DIAS"},
                {"value": "TRANSFERENCIA", "label": "TRANSFERENCIA"},
            ]
        return json_safe(out)


@app.get("/api/clientes")
def clientes(
    rut: str | None = Query(None),
    nombre: str | None = Query(None),
    ruta_id: str | None = Query(None),
    limit: int = Query(80, ge=1, le=500),
):
    filters = []
    params = []
    rut = clean_text(rut, 20)
    nombre = clean_text(nombre, 80)
    ruta_id = clean_text(ruta_id, 20)
    if rut:
        filters.append("AND c.cliente_rut LIKE %s")
        params.append(f"%{rut}%")
    if nombre:
        filters.append("AND c.cliente_nombre LIKE %s")
        params.append(f"%{nombre}%")
    if ruta_id:
        filters.append("AND CAST(c.ruta_id AS CHAR) = %s")
        params.append(ruta_id)

    where_extra = "\n".join(filters)
    with cursor() as (_, cur):
        cur.execute(f"""
            SELECT
                c.cliente_rut,
                c.cliente_nombre,
                c.cliente_direccion,
                c.Ciudad_codigo,
                c.Comuna,
                c.cliente_condiccion,
                COALESCE(fp.fpago_descripcion, c.cliente_condiccion) AS condicion_nombre,
                c.cliente_mail,
                c.cliente_estado,
                c.ruta_id,
                r.ruta_nombre,
                c.cliente_telefono,
                c.cliente_celular,
                c.cliente_giro,
                c.cliente_vendedor,
                c.lista_codigo,
                c.cliente_descuento,
                c.cliente_geo
            FROM clientes c
            LEFT JOIN rutas r ON r.ruta_id = c.ruta_id
            LEFT JOIN formasdepago fp ON fp.fpago_codigo = c.cliente_condiccion
            WHERE 1=1
              {where_extra}
            ORDER BY c.cliente_nombre, c.cliente_rut
            LIMIT %s
        """, tuple(params + [limit]))
        return json_safe(clean_rows(cur.fetchall()))


@app.get("/api/clientes/{cliente_rut}")
def cliente_detalle(cliente_rut: str):
    rut = clean_text(cliente_rut, 20)
    with cursor() as (_, cur):
        cur.execute("""
            SELECT *
            FROM clientes
            WHERE cliente_rut = %s
            LIMIT 1
        """, (rut,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Cliente no encontrado")
        return json_safe(clean_row(row))


@app.put("/api/clientes/{cliente_rut}")
def cliente_actualizar(cliente_rut: str, payload: SaveRequest):
    allowed = {
        "cliente_nombre", "cliente_direccion", "cliente_telefono", "cliente_celular",
        "cliente_mail", "cliente_giro", "Ciudad_codigo", "Comuna", "cliente_estado",
        "cliente_descuento", "ruta_id", "cliente_condiccion", "cliente_vendedor",
        "cliente_geo", "lista_codigo", "cliente_intercambio", "cliente_observacion",
        "cliente_banco", "cliente_factoring", "cliente_ctacte",
    }
    data = {key: value for key, value in payload.data.items() if key in allowed}
    if not data:
        raise HTTPException(400, "No hay datos para guardar")
    rut = clean_text(cliente_rut, 20)
    with cursor() as (_, cur):
        set_sql = ", ".join(f"`{key}` = %s" for key in data)
        cur.execute(f"UPDATE clientes SET {set_sql} WHERE cliente_rut = %s", tuple(data.values()) + (rut,))
        if cur.rowcount == 0:
            cur.execute("SELECT 1 FROM clientes WHERE cliente_rut = %s", (rut,))
            if not cur.fetchone():
                raise HTTPException(404, "Cliente no encontrado")
    return json_safe({"ok": True})


@app.get("/api/productos")
def productos(
    codigo: str | None = Query(None),
    descripcion: str | None = Query(None),
    familia: str | None = Query(None),
    proveedor: str | None = Query(None),
    lista: str = Query("01"),
    barra: str | None = Query(None),
    incluir_todos: bool = Query(False),
    limit: int = Query(80, ge=1, le=500),
):
    filters = []
    params = []
    codigo = clean_text(codigo, 30)
    descripcion = clean_text(descripcion, 100)
    familia = clean_text(familia, 30)
    proveedor = clean_text(proveedor, 30)
    lista = clean_text(lista, 5) or "01"
    barra = clean_text(barra, 20)
    if codigo:
        filters.append("AND p.producto_codigo LIKE %s")
        params.append(f"{codigo}%")
    if descripcion:
        filters.append("AND p.producto_descripcion LIKE %s")
        params.append(f"%{descripcion}%")
    if familia:
        filters.append("AND p.familia_codigo = %s")
        params.append(familia)
    if proveedor:
        filters.append("AND p.producto_proveedor = %s")
        params.append(proveedor)
    if barra:
        filters.append("AND p.producto_barra LIKE %s")
        params.append(f"%{barra}%")
    if not incluir_todos:
        filters.append("AND COALESCE(p.producto_estado, 'A') NOT IN ('I','N')")
    where_extra = "\n".join(filters)
    with cursor() as (_, cur):
        cur.execute(f"""
            SELECT
                p.producto_codigo,
                p.producto_descripcion,
                p.familia_codigo,
                p.unidad_codigo,
                p.producto_ubicacion,
                p.producto_costo,
                p.producto_venta,
                p.producto_estado,
                p.producto_barra,
                p.producto_unidadenvase,
                p.producto_margenvta,
                ROUND(COALESCE(p.producto_costo, 0) * (1 + COALESCE(p.producto_margenvta, 0) / 100), 0) AS producto_netoventa,
                COALESCE(st.producto_stock, 0) AS producto_stock,
                COALESCE(st.producto_stock, 0) * COALESCE(p.producto_costosinflete, 0) AS producto_valorizadocostosin,
                CASE WHEN COALESCE(p.producto_unidadenvase, 0) > 0
                     THEN p.producto_unidadenvase * COALESCE(p.producto_gramaje, 0)
                     ELSE 0 END AS producto_stockenvase,
                p.producto_descuento,
                p.producto_descuentastock,
                p.producto_stockmin,
                pr.proveedor_nombre,
                COALESCE(f.familia_descripcion, p.familia_codigo) AS familia_descripcion,
                COALESCE(pl.lista_neto, p.producto_neto, 0) AS lista_neto,
                COALESCE(pl.lista_iva, p.producto_iva, 0) AS lista_iva,
                COALESCE(pl.lista_venta, p.producto_venta, 0) AS lista_venta,
                COALESCE(pl.lista_margen, p.producto_margenvta, 0) AS lista_margen,
                COALESCE(pl.Lista_costo, p.producto_costo, 0) AS lista_costo
            FROM productos p
            LEFT JOIN proveedores pr ON pr.proveedor_codigo = p.producto_proveedor
            LEFT JOIN familias f ON f.familia_codigo = p.familia_codigo
            LEFT JOIN precioslevel1 pl ON pl.producto_codigo = p.producto_codigo AND pl.lista_codigo = %s
            LEFT JOIN (
                SELECT producto_codigo, SUM(COALESCE(producto_stockbodega, 0)) AS producto_stock
                FROM productoslevel2 GROUP BY producto_codigo
            ) st ON st.producto_codigo = p.producto_codigo
            WHERE 1=1
              {where_extra}
            ORDER BY p.producto_codigo, p.producto_descripcion
            LIMIT %s
        """, tuple([lista] + params + [limit]))
        return json_safe(clean_rows(cur.fetchall()))


@app.get("/api/catalogos/productos")
def catalogos_productos():
    queries = {
        "familias": "SELECT familia_codigo AS value, COALESCE(familia_descripcion, familia_codigo) AS label FROM familias ORDER BY familia_descripcion",
        "subfamilias": "SELECT familia_codigo AS parent, subfamilia_codigo AS value, COALESCE(Subfamilia_descripcion, subfamilia_codigo) AS label FROM familiaslevel1 ORDER BY familia_codigo, Subfamilia_descripcion",
        "proveedores": "SELECT proveedor_codigo AS value, COALESCE(proveedor_nombre, proveedor_codigo) AS label FROM proveedores ORDER BY proveedor_nombre",
        "unidades": "SELECT unidad_codigo AS value, COALESCE(unidad_descripcion, unidad_codigo) AS label FROM unidades ORDER BY unidad_descripcion",
        "listas": "SELECT lista_codigo AS value, COALESCE(lista_descripcion, lista_codigo) AS label FROM listaprecios ORDER BY lista_codigo",
        "impuestos": "SELECT impuesto_codigo AS value, CONCAT(impuesto_codigo, ' (', COALESCE(impuesto_valor, 0), '%)') AS label, COALESCE(impuesto_valor, 0) AS rate FROM impuestos ORDER BY impuesto_codigo",
    }
    out = {key: [] for key in queries}
    with cursor() as (_, cur):
        for key, sql in queries.items():
            cur.execute(sql)
            out[key] = clean_rows(cur.fetchall())
    out["estados"] = [{"value": "A", "label": "ACTIVO"}, {"value": "I", "label": "INACTIVO"}]
    out["si_no"] = [{"value": "S", "label": "SI"}, {"value": "N", "label": "NO"}]
    return json_safe(out)


@app.get("/api/productos/{producto_codigo}")
def producto_detalle(producto_codigo: str, lista: str = Query("01")):
    codigo = clean_text(producto_codigo, 20)
    lista = clean_text(lista, 5) or "01"
    with cursor() as (_, cur):
        cur.execute("""
            SELECT p.*, pr.proveedor_nombre, f.familia_descripcion,
                   pl.lista_neto, pl.lista_iva, pl.lista_venta, pl.lista_ila,
                   pl.lista_precio, pl.Lista_costo, pl.lista_margen
            FROM productos p
            LEFT JOIN proveedores pr ON pr.proveedor_codigo = p.producto_proveedor
            LEFT JOIN familias f ON f.familia_codigo = p.familia_codigo
            LEFT JOIN precioslevel1 pl ON pl.producto_codigo = p.producto_codigo AND pl.lista_codigo = %s
            WHERE p.producto_codigo = %s LIMIT 1
        """, (lista, codigo))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Producto no encontrado")
        product = clean_row(row)
        product.pop("producto_foto", None)
        product.pop("producto_doc", None)
        cur.execute("SELECT COALESCE(SUM(producto_stockbodega), 0) AS producto_stock FROM productoslevel2 WHERE producto_codigo = %s", (codigo,))
        stock = (cur.fetchone() or {}).get("producto_stock", 0)
        product.update(producto_virtual_values(product, stock))
        related_queries = {
            "comisiones": """
                SELECT l.empleado_codigo, COALESCE(e.empleado_nombre, l.empleado_codigo) AS empleado_nombre, l.producto_comision
                FROM productoslevel1 l LEFT JOIN empleados e ON e.empleado_codigo = l.empleado_codigo
                WHERE l.producto_codigo = %s ORDER BY empleado_nombre
            """,
            "bodegas": """
                SELECT l.bodega_codigo, COALESCE(b.bodega_descripcion, l.bodega_codigo) AS bodega_descripcion,
                       l.producto_stockbodega, l.producto_reservado
                FROM productoslevel2 l LEFT JOIN bodegas b ON b.bodega_codigo = l.bodega_codigo
                WHERE l.producto_codigo = %s ORDER BY l.bodega_codigo
            """,
            "ean": "SELECT producto_ean FROM productoslevel3 WHERE producto_codigo = %s ORDER BY producto_ean",
            "modelos": "SELECT marca_codigo, Modelo_codigo, aplicacion_desde, aplicacion_hasta FROM productoslevel4 WHERE producto_codigo = %s ORDER BY marca_codigo, Modelo_codigo",
            "compras": """
                SELECT l.producto_comprafecha, l.proveedor_codigo, COALESCE(p.proveedor_nombre, l.proveedor_codigo) AS proveedor_nombre,
                       l.producto_compracant, l.producto_compraprecio, l.producto_codigoproveedor,
                       l.producto_numerodoc, l.producto_vencimiento
                FROM productoslevel5 l LEFT JOIN proveedores p ON p.proveedor_codigo = l.proveedor_codigo
                WHERE l.producto_codigo = %s ORDER BY l.producto_comprafecha DESC LIMIT 100
            """,
            "cierres": "SELECT cierre_mes, cierre_ano, cierre_cantidad, cierre_venta, cierre_costoflete, cierre_costoneto, cierre_fecha FROM productoslevel6 WHERE producto_codigo = %s ORDER BY cierre_ano DESC, cierre_mes DESC",
            "componentes": """
                SELECT l.producto_codigo1, COALESCE(p.producto_descripcion, l.producto_codigo1) AS producto_descripcion,
                       l.producto_cantidad, l.producto_packprecio
                FROM productoslevel7 l LEFT JOIN productos p ON p.producto_codigo = l.producto_codigo1
                WHERE l.producto_codigo = %s ORDER BY producto_descripcion
            """,
            "aplicaciones": "SELECT producto_aplicacion FROM productoslevel8 WHERE producto_codigo = %s ORDER BY producto_aplicacion",
            "familias_relacionadas": """
                SELECT l.producto_familia, COALESCE(f.familia_descripcion, l.producto_familia) AS familia_descripcion
                FROM productoslevel9 l LEFT JOIN familias f ON f.familia_codigo = l.producto_familia
                WHERE l.producto_codigo = %s ORDER BY familia_descripcion
            """,
        }
        product["relaciones"] = {}
        for name, sql in related_queries.items():
            try:
                cur.execute(sql, (codigo,))
                product["relaciones"][name] = clean_rows(cur.fetchall())
            except Exception:
                product["relaciones"][name] = []
        return json_safe(product)


@app.put("/api/productos/{producto_codigo}")
def producto_actualizar(producto_codigo: str, payload: SaveRequest):
    product_allowed = {
        "producto_descripcion", "familia_codigo", "subfamilia_codigo", "producto_ubicacion",
        "producto_costo", "producto_costoof", "producto_venta", "producto_iva", "producto_ila",
        "producto_manejaiva", "producto_descuentastock", "impuesto_codigo", "producto_neto",
        "producto_stockmin", "unidad_codigo", "producto_unidadenvase", "producto_gramaje",
        "producto_proveedor", "producto_costoant", "producto_estado", "producto_peso",
        "producto_barra", "producto_margenvta", "producto_descuento", "producto_minimovta",
        "producto_oferta", "producto_ofertaneto", "producto_pack", "producto_serial",
        "producto_ventaant", "producto_costosinflete",
    }
    price_allowed = {"lista_neto", "lista_iva", "lista_venta", "lista_ila", "lista_precio", "Lista_costo", "lista_margen"}
    data = payload.data
    product_data = {key: value for key, value in data.items() if key in product_allowed}
    price_data = {key: value for key, value in data.items() if key in price_allowed}
    lista = clean_text(data.get("lista_codigo"), 5) or "01"
    codigo = clean_text(producto_codigo, 20)
    with cursor() as (_, cur):
        cur.execute("SELECT * FROM productos WHERE producto_codigo = %s LIMIT 1", (codigo,))
        current = clean_row(cur.fetchone() or {})
        if not current:
            raise HTTPException(404, "Producto no encontrado")
        calculated, impuesto_valor = producto_business_values(cur, codigo, product_data, current, False)
        product_data.update(calculated)
        costo = decimal_value(product_data.get("producto_costo", current.get("producto_costo")))
        neto_base = decimal_value(calculated["producto_neto"])
        price_data.update({
            "lista_neto": int(neto_base),
            "lista_iva": calculated["producto_iva"],
            "lista_ila": calculated["producto_ila"],
            "lista_venta": calculated["producto_venta"],
            "lista_precio": calculated["producto_venta"],
            "Lista_costo": float(costo),
            "lista_margen": float(((neto_base / costo) - 1) * 100) if costo else 0,
        })
        if product_data:
            set_sql = ", ".join(f"`{key}` = %s" for key in product_data)
            cur.execute(f"UPDATE productos SET {set_sql} WHERE producto_codigo = %s", tuple(product_data.values()) + (codigo,))
        if price_data:
            columns = ["lista_codigo", "producto_codigo"] + list(price_data)
            values = [lista, codigo] + list(price_data.values())
            updates = ", ".join(f"`{key}` = VALUES(`{key}`)" for key in price_data)
            cur.execute(
                f"INSERT INTO precioslevel1 ({', '.join(f'`{x}`' for x in columns)}) VALUES ({', '.join(['%s'] * len(columns))}) ON DUPLICATE KEY UPDATE {updates}",
                tuple(values),
            )
    return json_safe({"ok": True, "calculated": calculated, "impuesto_valor": impuesto_valor})


@app.post("/api/productos")
def producto_crear(payload: SaveRequest):
    data = dict(payload.data)
    codigo = clean_text(data.pop("producto_codigo", None), 20)
    if not codigo:
        raise HTTPException(400, "Ingrese el codigo del producto")
    lista = clean_text(data.pop("lista_codigo", None), 5) or "01"
    with cursor() as (_, cur):
        columns = table_columns(cur, "productos")
        product_data = {key: value for key, value in data.items() if key in columns and key not in {"producto_foto", "producto_doc"}}
        calculated, impuesto_valor = producto_business_values(cur, codigo, product_data, {}, True)
        product_data.update(calculated)
        product_data["producto_codigo"] = codigo
        keys = list(product_data)
        cur.execute(
            f"INSERT INTO productos ({', '.join(f'`{key}`' for key in keys)}) VALUES ({', '.join(['%s'] * len(keys))})",
            tuple(product_data[key] for key in keys),
        )
        costo = decimal_value(product_data.get("producto_costo"))
        neto_base = decimal_value(calculated["producto_neto"])
        price_data = {
            "lista_codigo": lista, "producto_codigo": codigo,
            "lista_neto": int(neto_base), "lista_iva": calculated["producto_iva"],
            "lista_ila": calculated["producto_ila"], "lista_venta": calculated["producto_venta"],
            "lista_precio": calculated["producto_venta"], "Lista_costo": float(costo),
            "lista_margen": float(((neto_base / costo) - 1) * 100) if costo else 0,
        }
        keys = list(price_data)
        cur.execute(
            f"INSERT INTO precioslevel1 ({', '.join(f'`{key}`' for key in keys)}) VALUES ({', '.join(['%s'] * len(keys))})",
            tuple(price_data[key] for key in keys),
        )
    return json_safe({"ok": True, "calculated": calculated, "impuesto_valor": impuesto_valor})


@app.get("/api/picking/resumen")
def picking_resumen(
    fecha: str = Query(..., description="YYYY-MM-DD"),
    ruta_id: str | None = Query(None),
    vendedor_codigo: str | None = Query(None),
):
    fecha_dt = parse_date(fecha)
    rows = picking_data(fecha_dt, clean_text(ruta_id, 20), clean_text(vendedor_codigo, 20))
    return json_safe({
        "fecha": fecha_dt.isoformat(),
        "productos": len(rows),
        "total_kilos": sum(to_float(r.get("kilos"), 0.0) for r in rows),
        "total_unidades": sum(to_float(r.get("unidades"), 0.0) for r in rows),
        "rows": rows,
    })


@app.get("/api/picking/pdf")
def picking_pdf(
    fecha: str = Query(..., description="YYYY-MM-DD"),
    ruta_id: str | None = Query(None),
    vendedor_codigo: str | None = Query(None),
):
    fecha_dt = parse_date(fecha)
    ruta = clean_text(ruta_id, 20)
    vendedor = clean_text(vendedor_codigo, 20)
    label_parts = []
    if ruta:
        label_parts.append(f"Ruta {ruta}")
    if vendedor:
        label_parts.append(f"Vendedor {vendedor}")
    filtro_label = " / ".join(label_parts) if label_parts else "Todas las rutas"
    rows = picking_data(fecha_dt, ruta, vendedor)
    pdf = picking_pdf_bytes(fecha_dt, filtro_label, rows)
    filename = f"picking_{fecha_dt.strftime('%Y%m%d')}.pdf"
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )
